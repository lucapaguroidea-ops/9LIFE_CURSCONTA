"""Asamblează dist/pachet/ — livrabilele de uz, ordonate după momentul în care ajungi
la ele, cu o pagină de index care le explică.

Pachetul e derivat integral din `dist/`. Nimic nu se copiază de mână și niciun număr nu
se scrie: titlurile vin din `date/documente.py`, numărătorile se citesc din workbook-uri
și din listele generate. Un index scris o dată ar rămâne în urmă la primul `make tot`,
cu numere false — exact al doilea adevăr pe care restul sistemului îl exclude.

Ordinea și judecata („când îl deschizi”, „la ce NU”) stau în `date/pachet.py`: singurul
lucru din pachet care nu se poate deduce.

Rulare:  python build/pachet.py   (după `make documente`)
"""
import os
import re
import shutil
import sys
import zipfile

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from build import html_out  # noqa: E402
from build import parcurs as bparc  # noqa: E402
from date import documente as ddoc  # noqa: E402
from date import intrebari as dintr  # noqa: E402
from date import pachet as P  # noqa: E402

RADACINA = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DIST = os.path.join(RADACINA, "dist")
PACHET = os.path.join(DIST, "pachet")
ARHIVA = os.path.join(DIST, "pachet-9life.zip")

PLAN = "Plan_de_conturi_ROL_Analitice_Fluxuri_SAGA.xlsx"
MODULE = "Module_Declarative_Fluxuri.xlsx"


def _doc(cheie):
    return next(c for c in ddoc.DOCUMENTE if c["cheie"] == cheie)


def surse(cheie):
    """Cheia de artefact → [(extensie, cale în dist/)]. Aici e singura hartă de fișiere."""
    if cheie == "xlsx:plan":
        return [(".xlsx", os.path.join(DIST, PLAN))]
    if cheie == "xlsx:module":
        return [(".xlsx", os.path.join(DIST, MODULE))]
    if cheie == "lista:intrebari":
        return [(".html", os.path.join(DIST, "intrebari-formator.html"))]
    if cheie == "lista:parcurs":
        return [(".html", os.path.join(DIST, "parcurs-training-nou.html"))]
    cfg = _doc(cheie)
    baza = os.path.join(RADACINA, cfg["iesire"])[:-3]
    return [(ext, baza + ext) for ext, _ in P.FORMATE]


def titlu(cheie):
    """Titlul afișat — din artefact, nu din `date/pachet.py`."""
    if cheie == "xlsx:plan":
        return "Plan de conturi — rol, analitice, fluxuri"
    if cheie == "xlsx:module":
        return "Module declarative de fluxuri"
    if cheie == "lista:intrebari":
        return "Întrebări pentru formator"
    if cheie == "lista:parcurs":
        return "Parcursul unui set nou de notițe"
    return _doc(cheie)["titlu"]


def cifre():
    """Numărătorile care descriu pachetul — citite, niciodată scrise.

    Dacă vreuna nu mai corespunde realității, pagina e veche, nu fișierele: se
    regenerează la fiecare `make tot`, odată cu ele.
    """
    wbp = openpyxl.load_workbook(os.path.join(DIST, PLAN), read_only=True)
    ws = wbp["Fluxuri"]
    fluxuri = set()
    conturi = set()
    for r in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        v = str(r[0] or "").strip()
        m = re.match(r"^(F-\d{3})\b", v)
        if m:
            fluxuri.add(m.group(1))
    for r in wbp["Plan de conturi"].iter_rows(min_col=1, max_col=1, values_only=True):
        v = str(r[0] or "").strip()
        if re.fullmatch(r"\d{3,4}([./]\d+)?", v):
            conturi.add(v)
    corelatii = {str(r[0] or "").strip()
                 for r in wbp["Corelații de control"].iter_rows(min_col=1, max_col=1,
                                                               values_only=True)
                 if re.fullmatch(r"C-\d\d", str(r[0] or "").strip())}
    foi_plan = len(wbp.sheetnames)
    wbp.close()

    wbm = openpyxl.load_workbook(os.path.join(DIST, MODULE), read_only=True)
    module = {str(r[0] or "").strip()
              for r in wbm["CatalogModule"].iter_rows(min_col=2, max_col=2,
                                                      values_only=True)
              if str(r[0] or "").strip().startswith("MOD_")}
    wbm.close()

    return dict(fluxuri=len(fluxuri), conturi=len(conturi), corelatii=len(corelatii),
                module=len(module), foi=foi_plan,
                intrebari=len(dintr.toate()), porti=len(bparc.porti()),
                documente=len(ddoc.DOCUMENTE))


def asambleaza():
    """Copiază livrabilele cu nume numerotate. Întoarce [(item, [(ext, nume)])]."""
    if os.path.isdir(PACHET):
        shutil.rmtree(PACHET)
    os.makedirs(PACHET)

    out = []
    for i, (etapa, cheie, nume, cand, nu) in enumerate(P.ITEME, start=1):
        fisiere = []
        for ext, cale in surse(cheie):
            if not os.path.exists(cale):
                raise SystemExit(f"pachet: lipsește {os.path.relpath(cale, RADACINA)} "
                                 f"— rulează `make build documente`")
            tinta = f"{i}-{nume}{ext}"
            shutil.copy2(cale, os.path.join(PACHET, tinta))
            fisiere.append((ext, tinta))
        out.append((etapa, cheie, nume, cand, nu, fisiere))
    return out


ETICHETA = {".xlsx": "Excel", ".docx": "Word", ".html": "de citit", ".md": "text"}
DESPRE_FORMAT = dict(P.FORMATE)


def index(iteme, n):
    """Pagina de index, în aceeași ținută grafică cu documentele."""
    L = [f"<p>{html_out._inline(P.INTRO)}</p>".replace("\n", " ")]

    for etapa, descriere in P.ETAPE:
        ale_etapei = [x for x in iteme if x[0] == etapa]
        if not ale_etapei:
            continue
        L.append(f'<h2 class="etapa">{html_out._inline(etapa)}</h2>')
        L.append(f'<p class="etapa-d">{html_out._inline(descriere)}</p>')
        for _, cheie, nume, cand, nu, fisiere in ale_etapei:
            nr = fisiere[0][1].split("-")[0]
            legaturi = " ".join(
                f'<a class="fmt" href="{f}">{ETICHETA.get(ext, ext)}'
                f'<span class="ext">{ext}</span></a>' for ext, f in fisiere)
            note = [DESPRE_FORMAT[ext] for ext, _ in fisiere if ext in DESPRE_FORMAT]
            L.append(
                f'<div class="item"><div class="item-cap">'
                f'<span class="item-nr">{nr}</span>'
                f'<span class="item-t">{html_out._inline(titlu(cheie))}</span></div>'
                f'<p class="cand">{html_out._inline(cand)}</p>'
                f'<p class="nu"><span class="nu-l">nu</span> '
                f'{html_out._inline(nu)}</p>'
                f'<div class="fmts">{legaturi}'
                + (f'<span class="fmt-n">{html_out._inline(" · ".join(note))}</span>'
                   if note else "")
                + '</div></div>')

    L.append('<h2 class="etapa">Ce e înăuntru, în cifre</h2>')
    L.append(html_out._tabel([
        "| Ce | Cât |",
        f"| Conturi în planul clasificat pe rol | {n['conturi']} |",
        f"| Fluxuri cu monografie pas cu pas | {n['fluxuri']} |",
        f"| Corelații de control | {n['corelatii']} |",
        f"| Module declarative executabile | {n['module']} |",
        f"| Foi în workbook-ul de referință | {n['foi']} |",
        f"| Documente de studiu | {n['documente']} |",
        f"| Întrebări rămase deschise | {n['intrebari']} |",
        f"| Porți de calitate trecute la generare | {n['porti']} |",
    ]))
    L.append(f'<p class="final">{html_out._inline(P.NOTA_FINAL)}</p>'.replace("\n", " "))

    corp = "\n".join(L)
    pagina = html_out.SABLON.format(
        titlu="Pachet de lucru — contabilitate",
        subtitlu="Livrabilele generate din depozit, în ordinea în care ajungi la ele.",
        css=html_out.CSS + CSS_PACHET, nav="", corp=corp)
    # pagina de index n-are cuprins lateral: are opt itemi, i-ai văzut pe toți deodată
    pagina = pagina.replace(
        '<nav class="rail" aria-label="Cuprins"><p class="rail-h">Cuprins</p><ol></ol></nav>',
        "")
    return pagina.replace('class="wrap"', 'class="wrap solo"')


CSS_PACHET = """
.wrap.solo{grid-template-columns:minmax(0,1fr)}
.wrap.solo main{max-width:74ch; margin:0 auto; padding-top:44px}
h2.etapa{font-size:1.45rem; margin:46px 0 6px}
p.etapa-d{color:var(--muted); font-size:.96rem; margin:0 0 20px; max-width:66ch}
.item{border:1px solid var(--rule); border-left:3px solid var(--accent);
  border-radius:0 8px 8px 0; background:var(--surface); padding:16px 20px 14px;
  margin:0 0 14px; box-shadow:var(--shadow)}
.item-cap{display:flex; gap:12px; align-items:baseline; margin-bottom:8px}
.item-nr{font-family:"IBM Plex Mono",ui-monospace,monospace; font-size:.78rem;
  font-variant-numeric:tabular-nums; color:var(--accent); background:var(--accent-soft);
  border:1px solid var(--accent-line); border-radius:5px; padding:2px 8px; flex:none}
.item-t{font-family:Spectral,Georgia,serif; font-weight:600; font-size:1.12rem;
  line-height:1.28; color:var(--ink)}
.item p{margin:0 0 10px; max-width:none}
p.cand{color:var(--ink-2); font-size:.95rem}
p.nu{color:var(--muted); font-size:.9rem}
.nu-l{font-family:"IBM Plex Mono",ui-monospace,monospace; font-size:.68rem;
  letter-spacing:.14em; text-transform:uppercase; color:var(--storno);
  border:1px solid var(--storno); border-radius:4px; padding:1px 6px; margin-right:7px}
.fmts{display:flex; gap:8px; align-items:center; flex-wrap:wrap; margin-top:12px}
a.fmt{display:inline-flex; gap:6px; align-items:baseline; text-decoration:none;
  font-family:"IBM Plex Mono",ui-monospace,monospace; font-size:.78rem;
  color:var(--accent); background:var(--accent-soft); border:1px solid var(--accent-line);
  border-radius:5px; padding:4px 10px; transition:background .12s}
a.fmt:hover{background:var(--surface)}
a.fmt:focus-visible{outline:2px solid var(--accent); outline-offset:1px}
.fmt .ext{color:var(--muted)}
.fmt-n{color:var(--muted); font-size:.82rem}
p.final{color:var(--muted); font-size:.9rem; border-top:1px solid var(--rule);
  padding-top:16px; margin-top:34px}
"""


def verifica_legaturi(cale_index):
    """Fiecare legătură din index trebuie să existe în folderul pachetului."""
    with open(cale_index, encoding="utf-8") as f:
        html = f.read()
    lipsa = [h for h in re.findall(r'href="([^"#:]+)"', html)
             if not os.path.exists(os.path.join(PACHET, h))]
    if lipsa:
        raise SystemExit(f"pachet: legături moarte în index — {lipsa}")
    return len(re.findall(r'class="fmt" href=', html))


def arhiveaza(cale_index):
    """Arhivează și verifică legăturile PE ARHIVĂ — aia se distribuie, nu folderul.

    Un index care navighează corect local, dar trimite în gol după dezarhivare, e mai
    rău decât niciunul: descoperi problema tocmai când n-ai repo-ul la îndemână.
    """
    with zipfile.ZipFile(ARHIVA, "w", zipfile.ZIP_DEFLATED) as z:
        for nume in sorted(os.listdir(PACHET)):
            z.write(os.path.join(PACHET, nume), os.path.join("pachet-9life", nume))

    with zipfile.ZipFile(ARHIVA) as z:
        inauntru = {os.path.basename(n) for n in z.namelist()}
        cu_index = z.read(os.path.join("pachet-9life",
                                       os.path.basename(cale_index))).decode("utf-8")
    lipsa = [h for h in re.findall(r'href="([^"#:]+)"', cu_index) if h not in inauntru]
    if lipsa:
        raise SystemExit(f"pachet: legături care nu există în arhivă — {lipsa}")
    return os.path.getsize(ARHIVA), len(inauntru)


def main():
    iteme = asambleaza()
    n = cifre()
    # nume fără diacritice: arhiva se dezarhivează și pe sisteme care mai
    # stâlcesc UTF-8 în numele de fișier, iar indexul e primul lucru deschis
    cale_index = os.path.join(PACHET, "0-CITESTE-INTAI.html")
    with open(cale_index, "w", encoding="utf-8") as f:
        f.write(index(iteme, n))
    legaturi = verifica_legaturi(cale_index)
    octeti, in_arhiva = arhiveaza(cale_index)
    print(f"scris: {os.path.relpath(PACHET, RADACINA)}/  "
          f"({len(iteme)} livrabile, {len(os.listdir(PACHET))} fișiere, "
          f"{legaturi} legături verificate)")
    print(f"scris: {os.path.relpath(ARHIVA, RADACINA)}  "
          f"({octeti / 1024:.0f} KB, {in_arhiva} fișiere, legături verificate în arhivă)")


if __name__ == "__main__":
    main()

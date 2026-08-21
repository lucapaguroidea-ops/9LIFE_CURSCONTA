"""Numărătorile care descriu sistemul — citite din artefacte, niciodată scrise.

Un singur loc, trei consumatori: indexul pachetului, blocul de cifre din `README.md`
și Legenda workbook-ului de plan. Dacă fiecare și-ar număra singur, cele trei ar începe
să difere — iar diferența nu s-ar vedea, pentru că toate ar arăta plauzibil.

Motivul pentru care fișierul ăsta există: README-ul chiar rămăsese în urmă. Afirma „23
corelații de control” când erau 29, și trimitea la `date/fluxuri.py`, fișier dispărut la
reorganizarea pe clase. Era singurul artefact cu cifre pe care nicio poartă nu-l
verifica. Apoi s-a văzut că nici Legenda nu era mai bună: „16 module declarative” când
erau 17, „cele 21 de întrebări rămase deschise” când erau 20, „81 conturi” când erau 80.

Cheia designului: `din_workbook()` numără dintr-un workbook **deschis**, nu din `dist/`.
Altfel e ou-și-găină — `build_plan.py` are nevoie de cifre ca să scrie Legenda, dar
`dist/` se scrie abia la final. Aceleași funcții servesc și README-ul, care deschide
fișierul gata scris. O singură logică de numărare, două momente în care se cheamă.
"""
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

RADACINA = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DIST = os.path.join(RADACINA, "dist")

PLAN = "Plan_de_conturi_ROL_Analitice_Fluxuri_SAGA.xlsx"
MODULE = "Module_Declarative_Fluxuri.xlsx"

RE_FLUX = re.compile(r"^(F-(\d)\d\d)\b")
RE_COREL = re.compile(r"^C-\d\d$")
RE_CONT = re.compile(r"^\d{3,4}([./]\d+)?$")
RE_MODUL = re.compile(r"^MOD_[A-Z_]+$")


def _prima_coloana(ws):
    for r in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        yield str(r[0] or "").strip()


def din_cod():
    """Cifrele care nu au nevoie de niciun workbook — se citesc din `date/` și `build/`."""
    from date import documente as ddoc
    from date import inchideri as dinch
    from date import intrebari as dintr
    from date import ordine as O
    from build import parcurs as bparc

    # Fluxurile se numără din `date/ordine.py`, nu din foaia `Fluxuri`. Motivul e de
    # moment, nu de principiu: `build_plan.py` cere cifrele ÎNAINTE de renumerotare,
    # iar atunci foaia încă poartă ID-urile vechi de două cifre (F-45), pe care
    # `RE_FLUX` nu le prinde — prima încercare a raportat 0 fluxuri. Numărul rămâne
    # verificat contra foii de poarta 10 („catalogul acoperă fix cele 68 monografii”).
    pe_clasa = {}
    for nou, _vechi, _den in O.ORDINE:
        pe_clasa.setdefault(nou[2], set()).add(nou)

    return dict(
        fluxuri=len(O.ORDINE),
        pe_clasa={c: len(v) for c, v in sorted(pe_clasa.items())},
        cadente=len(dinch.CADENTA),
        documente=len(ddoc.DOCUMENTE),
        intrebari=len(dintr.toate()),
        deschise=len([q for _, q in dintr.toate() if not q["raspuns"]]),
        verificate=len(dintr.verificate()),
        decizii=len(dintr.decizii()),
        porti=len(bparc.porti()),
    )


def din_workbook(wbp, wbm=None):
    """Numărătorile care se pot lua dintr-un workbook de plan DESCHIS.

    `wbp` poate fi workbook-ul care tocmai se construiește (în `build_plan.py`) sau
    unul încărcat din `dist/` (în README și pachet). Rezultatul e același, pentru că
    numără aceleași foi.

    `wbm` — workbook-ul de module, dacă e disponibil. `build_plan.py` nu-l are, și nici
    n-are nevoie: numărul de module îl citește din propria foaie `Index module`, care e
    generată tot din `date/module`. Aceeași sursă, alt drum.
    """
    conturi = {v for v in _prima_coloana(wbp["Plan de conturi"]) if RE_CONT.match(v)}
    corelatii = {v for v in _prima_coloana(wbp["Corelații de control"])
                 if RE_COREL.match(v)}
    rol_in_flux = {v for v in _prima_coloana(wbp["Doar rol în flux"]) if RE_CONT.match(v)}
    module = {v for v in _prima_coloana(wbp["Index module"]) if RE_MODUL.match(v)}

    # Două cifre diferite, ambele adevărate, cu etichete care spun ce numără:
    # câte conturi sunt CLASIFICATE Tier A în plan, și câte au rând DETALIAT în foaia
    # de analitice. README-ul afirma o a treia, care nu era niciuna din ele.
    ws = wbp["Plan de conturi"]
    cap = next(([str(x or "") for x in r]
                for r in ws.iter_rows(min_col=1, max_col=12, values_only=True)
                if str(r[0] or "").strip() == "Simbol"), None)
    i_tier = [k for k, x in enumerate(cap or []) if "Tier" in x]
    tier = {"A": 0, "B": 0, "C": 0}
    if i_tier:
        for r in ws.iter_rows(min_col=1, max_col=12, values_only=True):
            if not RE_CONT.match(str(r[0] or "").strip()):
                continue
            t = str(r[i_tier[0]] or "").strip().upper()[:1]
            if t in tier:
                tier[t] += 1
    tier_a = tier["A"]
    detaliate = len({v for v in _prima_coloana(wbp["Analitice (Tier A)"])
                     if RE_CONT.match(v)})

    n = dict(
        conturi=len(conturi), corelatii=len(corelatii),
        rol_in_flux=len(rol_in_flux), module=len(module),
        tier_a=tier_a, tier_b=tier["B"], tier_c=tier["C"], detaliate=detaliate,
        foi=len(wbp.sheetnames),
    )
    n.update(din_cod())
    if wbm is not None:
        n["foi_module"] = len(wbm.sheetnames)
    return n


def citeste():
    """Toate numărătorile, din workbook-urile scrise. Ridică dacă lipsesc."""
    cale_plan = os.path.join(DIST, PLAN)
    if not os.path.exists(cale_plan):
        raise SystemExit(f"cifre: lipsește {PLAN} — rulează `make build`")

    wbp = openpyxl.load_workbook(cale_plan, read_only=True)
    wbm = openpyxl.load_workbook(os.path.join(DIST, MODULE), read_only=True)
    try:
        return din_workbook(wbp, wbm)
    finally:
        wbp.close()
        wbm.close()


#: (tipar, cheie) — frazele cu cifră pe care generatoarele le emit efectiv.
#:
#: Poarta 24 verifică fiecare apariție a tiparului contra cifrei din `citeste()`.
#: Tiparele sunt ancorate în cuvintele din jur, nu generice, și motivul e important:
#: un `\d+ conturi` generic ar fi picat pe „Adăugate … 21 de conturi Tier A din clasele
#: 1 și 2”, care e o afirmație ISTORICĂ despre o adăugire, nu un total. Un semnal
#: zgomotos ar distruge încrederea în porți — aceeași judecată pentru care verificarea
#: totalurilor din proză n-a devenit niciodată poartă.
#:
#: Ce NU verifică poarta: un număr scris în proză pentru care nu există tipar aici.
#: Acoperirea e exact lista de mai jos, nici un rând mai mult.
#: `~?` nu e cosmetic. Forma din sămânță e „LISTA FLUXURILOR (~38)”, cu tildă, iar un
#: tipar care cere doar cifre n-o vede deloc — deci poarta trecea verde exact peste
#: cazul pe care trebuia să-l prindă. S-a văzut la testul negativ: am oprit aplicarea
#: pe foaia Fluxuri, „(~38)” a rămas în fișier, și poarta n-a clipit. O aproximare e
#: prin definiție altceva decât cifra derivată, deci trebuie să pice.
FRAZE = [
    (r"LISTA FLUXURILOR \(~?(\d+)\)", "fluxuri"),
    (r"~?(\d+) fluxuri × pași", "fluxuri"),
    (r"Fluxuri: ~?(\d+)/\d+ detaliate", "fluxuri"),
    (r"~?(\d+) module declarative", "module"),
    (r"cele ~?(\d+) întrebări rămase deschise", "deschise"),
    (r"Cele ~?(\d+) documente de studiu", "documente"),
    (r"~?(\d+) conturi, cu 3 coloane noi", "conturi"),
    (r"~?(\d+) conturi de serviciu", "rol_in_flux"),
    (r"Tier A \(~?(\d+)", "tier_a"),
    (r"Tier B \(~?(\d+)\)", "tier_b"),
    (r"Tier C \(~?(\d+)\)", "tier_c"),
]


#: Foi în care cifrele VECHI au voie să rămână, pentru că exact ăsta le e rostul.
#: `Istoric` păstrează verbatim ce spunea fișierul înainte de curățare — „Fluxuri: 38/38
#: detaliate”, „81 conturi de serviciu”. Ar fi absurd s-o cerem la zi: e memoria, nu
#: foaia de lucru. Fără excepția asta, poarta ar cere ștergerea propriei evidențe.
FOI_ISTORICE = {"Istoric"}


def fraze_gresite(cai, n=None):
    """[(fișier, foaie, celulă, fraza, găsit, așteptat)] — frazele cu cifră greșită."""
    n = n or citeste()
    tipare = [(re.compile(t), k) for t, k in FRAZE]
    out = []
    for cale in cai:
        wb = openpyxl.load_workbook(cale, read_only=True)
        try:
            for ws in wb.worksheets:
                if ws.title in FOI_ISTORICE:
                    continue
                for row in ws.iter_rows():
                    for cell in row:
                        v = cell.value
                        if not isinstance(v, str):
                            continue
                        for rx, cheie in tipare:
                            for m in rx.finditer(v):
                                if int(m.group(1)) != n[cheie]:
                                    out.append((os.path.basename(cale), ws.title,
                                                cell.coordinate, m.group(0),
                                                int(m.group(1)), n[cheie]))
        finally:
            wb.close()
    return out

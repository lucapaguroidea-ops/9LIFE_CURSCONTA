"""Porțile de calitate ale sistemului. Cod de ieșire ≠ 0 la orice eșec.

Lista de mai jos e SURSA CANONICĂ a descrierilor: documentul de parcurs o citește de
aici, iar poarta 15 verifică faptul că numerele din listă coincid cu porțile chemate
efectiv mai jos. Fără verificarea aia, lista ar rămâne în urmă — s-a și întâmplat: a
descris multă vreme nouă porți, cu poarta 9 în forma ei veche.

  1. ΣDebit = ΣCredit pe fiecare pas de flux care are sume
  2. fiecare flux se închide cu un pas de verificare, o stare terminală și „Principiul:”
  3. fiecare flux didactic ★ are exact un pas revelator
  4. matricea nu are goluri nedeclarate, iar fiecare flux nou e referit în ea
  5. fiecare analitic Tier A are un factor din D/N/C/F/B/V/O și spune ce se rupe fără el
  6. fiecare token MOD_* referit există în CatalogModule (verificare între fișiere)
  7. corelațiile se verifică pe cifrele fluxurilor, nu declarativ
  8. formule echilibrate, niciun text scris din greșeală ca formulă, zero erori după
     recalc, toate celulele Check = OK
  9. conservare: fiecare linie din workbook-urile originale se regăsește în cele
     generate; înlocuirile intenționate se declară în `date/reformulari.py`, cu motiv
 10. catalogul de fluxuri acoperă fix monografiile — nici mai mult, nici mai puțin
 11. zero nume definite rupte
 12. conservare pe documentele revizuite: nicio linie pierdută la armonizare
 13. documentele au aceeași legendă, anexe canonice în ordine, și toate trei
     formatele: .md, .docx, .html
 14. tabelul de structură din foaia Legendă cunoaște toate foile workbook-ului
 15. documentul de parcurs nu citează foi, fișiere sau porți care nu există
 16. o sursă împărțită pe mai multe destinații nu pierde nimic în cusătură:
     fiecare subsecțiune are destinație declarată și ajunge exact acolo
 17. disciplina de închidere e ancorată în ambele sensuri: fiecare cont urmărit
     periodic e starea terminală a unui flux, iar fiecare cont cu rol în flux care
     se golește are o cadență — sau un motiv declarat pentru care nu are
 18. monografiile scrise în proză se echilibrează, iar aritmetica afirmată în text
     („5% din 250 = 12,50”) chiar se verifică — acolo unde poarta 1 nu ajunge
 19. marcajul ❓ e aplicat, nu doar definit: un document pe care o întrebare deschisă
     îl privește îl poartă, iar un document care îl poartă are întrebări deschise
 20. fiecare cont folosit într-un pas de flux există în „Plan de conturi”
 21. fiecare flux declarat de un modul în `CATALOG['fluxuri']` există cu adevărat
 22. blocul de cifre din README e exact cel pe care generatorul l-ar produce
 23. foile pe care catalogul le numește există, iar fiecare foaie de modul are exact
     o intrare de catalog — în ambele sensuri
 25. nicio foaie de modul nu mai vine din sămânță: fiecare are un generator în `date/`
 26. harta de diacritice e aplicată complet: niciun cuvânt din partea ei stângă nu mai
     apare în foile pe care le acoperă, iar simbolurile de cont sunt neatinse
 24. nicio frază cu cifră din workbook-uri nu contrazice cifra reală („17 module
     declarative”, „68 fluxuri × pași”) — foaia Istoric e scutită, acolo cifrele
     vechi sunt chiar conținutul

Rulare:  python build/verifica.py
"""
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from date import ANALITICE, CORELATII, FLUXURI, plan as dplan  # noqa: E402
from date import module as dmod  # noqa: E402
from date import ordine as O, reformulari as dreform  # noqa: E402
from date import repartizare as drep  # noqa: E402
from date import inchideri as dinch  # noqa: E402
from date import monografii as dmono  # noqa: E402
from date import intrebari as dintr  # noqa: E402
from build import conservare  # noqa: E402
from build import documente as bdoc  # noqa: E402
from build import repartizare as brep  # noqa: E402
from build import inchideri as binch  # noqa: E402
from build import monografii as bmono  # noqa: E402
from build import cifre as bcifre  # noqa: E402
from date.module import comun as dmcomun  # noqa: E402
from date import diacritice as ddiac  # noqa: E402
from build import diacritice as bdiac  # noqa: E402
from build import readme as breadme  # noqa: E402
from date import documente as ddoc  # noqa: E402

RADACINA = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SURSE = os.path.join(RADACINA, "surse", "training-4-2026-08-14")
DIST = os.path.join(RADACINA, "dist")
PLAN = os.path.join(DIST, "Plan_de_conturi_ROL_Analitice_Fluxuri_SAGA.xlsx")
MODULE = os.path.join(DIST, "Module_Declarative_Fluxuri.xlsx")

FACTORI = set("DNCFBVO")
ERORI_EXCEL = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!")

esecuri = []
note = []


def _echilibru(formula):
    """(adâncime_paranteze, șir_rămas_deschis) pentru o formulă Excel.

    Parantezele din interiorul șirurilor nu se numără, iar `""` este ghilimeaua
    scăpată din Excel, nu un sfârșit de șir.
    """
    adanc, in_sir, i = 0, False, 0
    while i < len(formula):
        ch = formula[i]
        if ch == '"':
            if in_sir and i + 1 < len(formula) and formula[i + 1] == '"':
                i += 2
                continue
            in_sir = not in_sir
        elif not in_sir:
            if ch == "(":
                adanc += 1
            elif ch == ")":
                adanc -= 1
        i += 1
    return adanc, in_sir


def cade(poarta, mesaj):
    esecuri.append(f"[{poarta}] {mesaj}")


def ok(poarta, mesaj):
    note.append(f"  ✔ {poarta}: {mesaj}")


# ---------------------------------------------------------------- porțile 1, 2, 3
def poarta_fluxuri():
    dezechilibrate = 0
    for f in FLUXURI:
        pasi_cu_sume = 0
        revelatori = 0
        for p in f["pasi"]:
            if p["revelator"]:
                revelatori += 1
            if not p["dr"] and not p["cr"]:
                continue
            pasi_cu_sume += 1
            sd = round(sum(s for _, s in p["dr"]), 2)
            sc = round(sum(s for _, s in p["cr"]), 2)
            if abs(sd - sc) > 0.005:
                cade("1", f"{f['id']} pas {p['nr']}: ΣD={sd} ≠ ΣC={sc}")
                dezechilibrate += 1
            if not p["dr"] or not p["cr"]:
                cade("1", f"{f['id']} pas {p['nr']}: are sume doar pe o parte")

        # poarta 2 — ultimul pas trebuie să fie verificare cu stare terminală
        ultim = f["pasi"][-1]
        rol = (ultim["rol"] or "").lower()
        if "stare terminală" not in rol and "stare terminala" not in rol:
            cade("2", f"{f['id']}: ultimul pas nu declară o stare terminală (rol={ultim['rol']!r})")
        if ultim["dr"] or ultim["cr"]:
            cade("2", f"{f['id']}: ultimul pas are sume — ar trebui să fie pas de verificare")

        # poarta 3 — fluxurile didactice au exact un pas revelator
        if f["didactic"].startswith("★"):
            if revelatori != 1:
                cade("3", f"{f['id']} (didactic): are {revelatori} pași revelatori, se cere exact 1")
        elif revelatori > 1:
            cade("3", f"{f['id']}: are {revelatori} pași revelatori pe un flux nedidactic")

        if not f["principiu"]:
            cade("2", f"{f['id']}: nu are „Principiul:” la finalul blocului")

    if not esecuri:
        ok("1", f"ΣD = ΣC pe toți pașii cu sume din {len(FLUXURI)} fluxuri")
        ok("2", "toate fluxurile se închid cu stare terminală și principiu")
        ok("3", "fluxurile didactice au exact un pas revelator")
    return dezechilibrate


# ------------------------------------------------------------------------ poarta 5
def poarta_factori():
    for a in ANALITICE:
        f = (a["factor"] or "").replace(" ", "")
        if not f or f == "—":
            cade("5", f"analitic {a['simbol']}: fără factor")
        elif not set(f) <= FACTORI:
            cade("5", f"analitic {a['simbol']}: factor necunoscut {a['factor']!r}")
        if not a["rupe"]:
            cade("5", f"analitic {a['simbol']}: nu spune ce se rupe dacă lipsește")
    ok("5", f"toate cele {len(ANALITICE)} analitice Tier A au factor din D/N/C/F/B/V/O")


# ------------------------------------------------------------------------ poarta 4
def poarta_acoperire(wb):
    # fluxurile din date/ poartă ID-uri vechi; matricea le are pe cele noi
    ids = {O.HARTA.get(f["id"], f["id"]) for f in FLUXURI}
    ws = wb["Matrice acoperire"]
    goluri = []
    referite = set()
    for r in range(1, ws.max_row + 1):
        gol = ws.cell(row=r, column=6).value
        simbol = ws.cell(row=r, column=1).value
        fluxuri = ws.cell(row=r, column=4).value
        if simbol and gol and str(gol).strip() not in ("Gol?", "NU"):
            s = str(simbol).strip()
            if s not in dplan.GOLURI_ACCEPTATE:
                goluri.append(f"{simbol} = {gol}")
        if fluxuri:
            referite |= set(re.findall(r"F-\d+", str(fluxuri)))
    if goluri:
        cade("4", "goluri nedeclarate în matrice: " + "; ".join(goluri))
    else:
        if dplan.GOLURI_ACCEPTATE:
            ok("4", f"fără goluri nedeclarate; {len(dplan.GOLURI_ACCEPTATE)} rămân marcate "
                    f"onest: {', '.join(sorted(dplan.GOLURI_ACCEPTATE))}")
        else:
            ok("4", "matricea nu are niciun gol — toate conturile Tier A sunt acoperite")

    # marcajele pe care extinderea s-a angajat să le rezolve trebuie chiar rezolvate
    for s in dplan.PARTIAL_REZOLVATE:
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(row=r, column=1).value or "").strip() == s:
                if str(ws.cell(row=r, column=6).value or "").strip() != "NU":
                    cade("4", f"{s}: promis rezolvat, dar marcajul nu e „NU”")
                break

    nereferite = ids - referite
    if nereferite:
        cade("4", f"fluxuri noi neapărute în matrice: {sorted(nereferite)}")
    else:
        ok("4", f"toate cele {len(ids)} fluxuri noi sunt referite în matrice")


# ------------------------------------------------------------------------ poarta 7
def poarta_corelatii():
    """Corelațiile se verifică pe cifrele fluxurilor, nu declarativ."""
    ids = {f["id"] for f in FLUXURI}
    sume = {}
    for f in FLUXURI:
        for p in f["pasi"]:
            for cont, s in p["dr"]:
                sume.setdefault((f["id"], "D", cont.split(".")[0]), 0)
                sume[(f["id"], "D", cont.split(".")[0])] += s
            for cont, s in p["cr"]:
                sume.setdefault((f["id"], "C", cont.split(".")[0]), 0)
                sume[(f["id"], "C", cont.split(".")[0])] += s

    def rulaj(flux, sens, cont):
        return round(sume.get((flux, sens, cont), 0), 2)

    # C-13 pe F-59: amortizarea descărcată nu poate depăși valoarea de intrare
    if rulaj("F-59", "D", "2812") > rulaj("F-59", "C", "212"):
        cade("7", "C-13 pe F-59: amortizarea descărcată depășește valoarea activului")

    # F-59: 212 se stinge integral (30.000 + 70.000 = 100.000)
    if rulaj("F-59", "C", "212") != 100000:
        cade("7", f"F-59: 212 nu se stinge integral (credit {rulaj('F-59', 'C', '212')})")

    # F-50: valoarea de intrare a mijlocului fix = 150.000 + 5.250 = 155.250
    val_2133 = rulaj("F-50", "D", "2133")
    if val_2133 != 155250:
        cade("7", f"F-50: valoarea de intrare 2133 = {val_2133}, se aștepta 155.250")
    # F-50: 4093 se deschide și se închide integral
    if rulaj("F-50", "D", "4093") != rulaj("F-50", "C", "4093"):
        cade("7", "F-50: avansul 4093 nu se închide integral în 167")
    # F-50: amortizarea contabilă = 155.250 / 60
    amo = rulaj("F-50", "D", "6811")
    if abs(amo - round(155250 / 60, 2)) > 0.01:
        cade("7", f"F-50: amortizarea lunară {amo} ≠ 155.250/60")

    # F-49: soldul în lei după reevaluare = sold valută × curs BNR
    # 20.000 EUR @4,97 = 99.400 iniţial; rată 5.000 @4,97 = 24.850; reevaluare +600
    sold_1621 = 99400 - rulaj("F-49", "D", "1621") + rulaj("F-49", "C", "1621")
    if abs(sold_1621 - 15000 * 5.01) > 0.01:
        cade("7", f"F-49: sold 1621 = {sold_1621}, se aștepta 15.000 × 5,0100 = 75.150")

    # F-46: repartizarea acoperă integral profitul de 2.500
    if rulaj("F-46", "D", "121") != 2500:
        cade("7", f"F-46: 121 nu se închide integral (debit {rulaj('F-46', 'D', '121')})")
    # F-46: rezerva legală = 5% din 2.500
    if rulaj("F-46", "C", "1061") != 125:
        cade("7", "F-46: rezerva legală ≠ 5% din profitul contabil brut")

    # F-47 / F-51: conturile tranzitorii se închid integral
    for flux, cont in (("F-47", "1174"), ("F-51", "1511"), ("F-52", "233"),
                       ("F-58", "231"), ("F-57", "223"), ("F-45", "456")):
        d, c = rulaj(flux, "D", cont), rulaj(flux, "C", cont)
        if d != c:
            cade("7", f"{flux}: contul {cont} nu se închide (D={d}, C={c})")

    # F-52 / F-58: capitalizarea neutralizează integral cheltuiala cu salariile
    for flux, venit in (("F-52", "721"), ("F-58", "722")):
        if rulaj(flux, "C", venit) != rulaj(flux, "D", "641"):
            cade("7", f"{flux}: {venit} nu neutralizează integral cheltuiala 641")

    if not any(e.startswith("[7]") for e in esecuri):
        ok("7", f"corelațiile se verifică pe cifrele din {len(ids)} fluxuri")


# --------------------------------------------------------------------- porțile 6, 8
def poarta_module():
    if not os.path.exists(MODULE):
        note.append("  – 6/8: Module_Declarative_Fluxuri.xlsx încă negenerat, porți sărite")
        return
    wb = openpyxl.load_workbook(MODULE)
    catalog = set()
    ws = wb["CatalogModule"]
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v and str(v).startswith("MOD_"):
            catalog.add(str(v).strip())

    referite = set()
    wbp = openpyxl.load_workbook(PLAN)
    for foaie in ("Corelații de control", "Index module"):
        w = wbp[foaie]
        for row in w.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    referite |= set(re.findall(r"MOD_[A-Z_]+", cell))
    lipsa = referite - catalog
    if lipsa:
        cade("6", f"module referite dar absente din CatalogModule: {sorted(lipsa)}")
    else:
        ok("6", f"toate cele {len(referite)} module referite există în CatalogModule")

    # poarta 8a — formule sintactic valide (paranteze și ghilimele echilibrate).
    # Fără asta, o formulă stricată e prinsă abia de recalc, cu un traceback opac.
    stricate = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    adanc, in_sir = _echilibru(c.value)
                    if adanc != 0 or in_sir:
                        cade("8", f"{ws.title}!{c.coordinate}: formulă dezechilibrată "
                                  f"(paranteze={adanc}, șir deschis={in_sir})")
                        stricate += 1
    if stricate == 0:
        ok("8", "toate formulele au paranteze și ghilimele echilibrate")

    # poarta 8c — proză scrisă din greșeală ca formulă.
    # `openpyxl` scrie orice text care începe cu „=” drept formulă, iar Excel afișează
    # #NAME?. Motorul de recalc nici măcar nu semnalează: nu poate parsa, deci nu
    # injectează valoare, deci celula rămâne tăcut stricată. Semnul distinctiv e spațiul
    # imediat după egal — nicio formulă reală nu arată așa.
    #
    # A doua formă, găsită la portare: text care arată ca o formulă PERFECT VALIDĂ, dar
    # referă o foaie care nu există în workbook-ul ăsta. Foaia `Istoric` citează
    # înlocuirile declarate, iar odată cu modulele portate au ajuns acolo formule
    # „=IF(CatalogModule!A13=…)”. În workbook-ul de plan nu există CatalogModule, deci
    # celula era stricată tăcut — recalcul se plângea, build-ul mergea mai departe.
    # De-asta se verifică și că fiecare foaie referită de o formulă chiar există.
    proza = 0
    for cale in (PLAN, MODULE):
        wbx = openpyxl.load_workbook(cale)
        foi_reale = set(wbx.sheetnames)
        for ws in wbx.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.data_type != "f" or not isinstance(c.value, str):
                        continue
                    for foaie in set(re.findall(r"([A-Za-zĂÂÎȘȚăâîșț_][\w ĂÂÎȘȚăâîșț.]*)!",
                                                c.value)):
                        if foaie.strip("'") not in foi_reale:
                            cade("8", f"{os.path.basename(cale)} {ws.title}!"
                                      f"{c.coordinate}: formula referă foaia "
                                      f"{foaie!r}, care nu există")
                            proza += 1
        for ws in openpyxl.load_workbook(cale).worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and re.match(r"^=\s", c.value):
                        cade("8", f"{os.path.basename(cale)} {ws.title}!{c.coordinate}: "
                                  f"text scris ca formulă — {c.value[:40]!r}")
                        proza += 1
    if proza == 0:
        ok("8", "nicio celulă de text nu e scrisă din greșeală ca formulă")

    # poarta 8b — valori de formulă
    wbv = openpyxl.load_workbook(MODULE, data_only=True)
    erori, checkuri, checkuri_ok = 0, 0, 0
    for ws in wbv.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str):
                    if v.strip() in ERORI_EXCEL:
                        cade("8", f"{ws.title}!{c.coordinate}: {v}")
                        erori += 1
                    if v.startswith("OK") or v.startswith("EROARE"):
                        checkuri += 1
                        if v.startswith("OK"):
                            checkuri_ok += 1
                        else:
                            cade("8", f"{ws.title}!{c.coordinate}: {v}")
    if erori == 0:
        ok("8", f"zero erori de formulă; {checkuri_ok}/{checkuri} celule Check = OK")


# ------------------------------------------------------------- porțile 9, 10, 11
def poarta_conservare():
    """Nimic din conținutul original nu dispare — verificat ca MULȚIME, nu ca ordine.

    Înlocuiește vechea poartă de „rânduri originale în aceeași ordine”, care interzicea
    tocmai reordonarea cerută. Renumerotarea nu contează ca pierdere: textul original e
    trecut prin harta F-vechi → F-nou înainte de căutare.
    """
    surse = [os.path.join(SURSE, n) for n in
             ("Plan_de_conturi_ROL_Analitice_Fluxuri_SAGA.xlsx",
              "Module_Declarative_Fluxuri.xlsx")]
    pierdute = conservare.verifica(surse, [PLAN, MODULE], O.HARTA,
                                   set(dreform.DECLARATE))
    if pierdute:
        for t_ in pierdute[:10]:
            cade("9", f"text original pierdut nedeclarat: {t_[:100]}")
        if len(pierdute) > 10:
            cade("9", f"…și încă {len(pierdute) - 10} linii")
    else:
        ok("9", f"conservare: tot textul original se regăsește "
                f"({len(dreform.DECLARATE)} înlocuiri declarate, cu motiv)")


def poarta_catalog(wb):
    """Catalogul trebuie să acopere fix monografiile — nici mai mult, nici mai puțin.

    În originalul training 4, 13 din 44 de fluxuri lipseau din catalog. Catalogul e
    acum derivat din monografii, deci poarta asta ar trebui să fie imposibil de rupt —
    exact de aceea merită verificată.
    """
    ws = wb["Fluxuri"]
    catalog, monografii = set(), set()
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(row=r, column=1).value or "").strip()
        d = str(ws.cell(row=r, column=4).value or "").strip()
        if re.fullmatch(r"F-\d{3}", a) and d in ("nu", "★ DA"):
            catalog.add(a)
        m = re.match(r"^(F-\d{3})\s+—", a)
        if m:
            monografii.add(m.group(1))
    if catalog - monografii:
        cade("10", f"în catalog dar fără monografie: {sorted(catalog - monografii)}")
    if monografii - catalog:
        cade("10", f"cu monografie dar absente din catalog: {sorted(monografii - catalog)}")
    if catalog == monografii:
        ok("10", f"catalogul acoperă fix cele {len(catalog)} monografii")


def poarta_nume_definite():
    """Niciun nume definit rupt.

    Templateul extern de deconturi are `decl_nr_decont` → `#REF!`. E exact defectul pe
    care nu vrem să-l reproducem când trecem modulele pe nume definite.
    """
    for cale in (PLAN, MODULE):
        wbn = openpyxl.load_workbook(cale)
        rupte = {n: str(d.value) for n, d in wbn.defined_names.items()
                 if "#REF" in str(d.value)}
        if rupte:
            cade("11", f"{os.path.basename(cale)}: nume definite rupte: {rupte}")
    if not any(e.startswith("[11]") for e in esecuri):
        ok("11", "zero nume definite rupte")


def poarta_structura(wb):
    """Poarta 14 — tabelul de structură din Legendă știe de toate foile.

    Tabelul enumera șase foi dintr-un workbook care ajunsese la unsprezece. E exact
    genul de listă scrisă de mână care rămâne în urmă fără să spună nimeni; verificată,
    nu mai poate.
    """
    ws = wb["Legendă"]
    listate = set()
    inauntru = False
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(row=r, column=1).value or "").strip()
        if a.startswith("5. STRUCTURA"):
            inauntru = True
            continue
        if inauntru and a.startswith("6. "):
            break
        if inauntru and a and a != "Foaie":
            listate.add(a)
    lipsa = [s for s in wb.sheetnames if s not in listate]
    if lipsa:
        cade("14", f"foi absente din tabelul de structură al Legendei: {lipsa}")
    else:
        ok("14", f"tabelul de structură cunoaște toate cele {len(wb.sheetnames)} foi")


def poarta_parcurs(wb):
    """Poarta 15 — documentul de parcurs nu minte despre starea sistemului.

    Două lucruri, ambele fiind aceeași idee: o listă scrisă de mână rămâne în urmă dacă
    nimeni nu o verifică.

    a) Lista canonică de porți din docstring-ul acestui fișier trebuie să coincidă cu
       porțile chemate efectiv. A divergit deja o dată: descria nouă porți, cu poarta 9
       în forma ei veche.
    b) Documentul de parcurs nu citează fișiere, module, fluxuri sau porți inexistente.
    """
    import build.parcurs as bpar

    # (a) lista canonică vs. porțile chemate
    listate = {n for n, _ in bpar.porti()}
    with open(os.path.abspath(__file__), encoding="utf-8") as f:
        corp = f.read().split('"""', 2)[2]
    chemate = {int(n) for n in re.findall(r'(?:ok|cade)\(\s*"(\d+)"', corp)}
    if listate != chemate:
        if chemate - listate:
            cade("15", f"porți chemate dar absente din lista canonică: "
                       f"{sorted(chemate - listate)}")
        if listate - chemate:
            cade("15", f"porți în lista canonică dar nechemate niciodată: "
                       f"{sorted(listate - chemate)}")
    else:
        ok("15", f"lista canonică de porți coincide cu cele {len(chemate)} chemate")

    # (b) referințele documentului
    cale = os.path.join(DIST, "parcurs-training-nou.md")
    if not os.path.exists(cale):
        cade("15", "lipsește dist/parcurs-training-nou.md — rulează `make parcurs`")
        return
    with open(cale, encoding="utf-8") as f:
        text = f.read()

    lipsa = []
    for fis in set(re.findall(r"`((?:date|build|surse|dist)/[\w./-]+)`", text)):
        if not os.path.exists(os.path.join(RADACINA, fis)):
            lipsa.append(f"fișier inexistent: {fis}")

    wm = openpyxl.load_workbook(MODULE)["CatalogModule"]
    module_reale = {str(wm.cell(row=r, column=2).value or "").strip()
                    for r in range(1, wm.max_row + 1)}
    for cod in set(re.findall(r"`(MOD_[A-Z_]+)`", text)):
        if cod not in module_reale:
            lipsa.append(f"modul inexistent: {cod}")

    for nr in set(re.findall(r"poarta (\d+)", text)):
        if int(nr) not in listate:
            lipsa.append(f"poartă inexistentă: {nr}")

    fluxuri_reale = set()
    wf = wb["Fluxuri"]
    for r in range(1, wf.max_row + 1):
        a = str(wf.cell(row=r, column=1).value or "").strip()
        if re.fullmatch(r"F-\d{3}", a):
            fluxuri_reale.add(a)
    for fid in set(re.findall(r"`(F-\d{3})`", text)):
        # ID-urile „următorul liber” sunt intenționat inexistente încă
        if fid not in fluxuri_reale and fid not in {O.urmatorul_liber(c)
                                                    for c, _, _ in O.BLOCURI}:
            lipsa.append(f"flux inexistent: {fid}")

    for l in lipsa:
        cade("15", f"documentul de parcurs citează {l}")
    if not lipsa:
        ok("15", "documentul de parcurs nu citează nimic inexistent")


def poarta_documente():
    """Porțile 12 și 13 — pe documentele revizuite, nu pe workbook-uri.

    12: aceeași garanție ca la Excel-uri — nicio linie din documentul original nu
        dispare la armonizare.
    13: cele trei documente arată ca același gen de document: aceeași legendă, anexe
        denumite canonic, în ordine.
    """
    pierdute_total = 0
    for cfg in ddoc.DOCUMENTE:
        cale = os.path.join(RADACINA, cfg["iesire"])
        if not os.path.exists(cale):
            cade("12", f"{cfg['nume']}: lipsește {cfg['iesire']} — rulează `make documente`")
            continue
        nou_text, original = bdoc.armonizeaza(cfg)
        lipsa = bdoc.verifica_conservare(original, nou_text, cfg)
        if lipsa:
            pierdute_total += len(lipsa)
            for l in lipsa[:5]:
                cade("12", f"{cfg['nume']}: linie pierdută — {l[:90]}")
    if pierdute_total == 0 and not any(e.startswith("[12]") for e in esecuri):
        ok("12", f"conservare pe cele {len(ddoc.DOCUMENTE)} documente revizuite: "
                 f"zero linii pierdute la armonizare")

    # 13 — aceeași legendă și anexe canonice
    titlu_legenda = f"## {ddoc.LEGENDA_TITLU}"
    for cfg in ddoc.DOCUMENTE:
        cale = os.path.join(RADACINA, cfg["iesire"])
        if not os.path.exists(cale):
            continue
        with open(cale, encoding="utf-8") as f:
            text = f.read()
        if text.count(titlu_legenda) != 1:
            cade("13", f"{cfg['nume']}: legenda apare de {text.count(titlu_legenda)} ori, "
                       f"se cere exact o dată")
        for m, _ in ddoc.LEGENDA_TABEL:
            if f"| {m} |" not in text:
                cade("13", f"{cfg['nume']}: lipsește marcajul {m} din legendă")
        gasite = re.findall(r"^## Anexa ([A-G]) — (.+)$", text, flags=re.M)
        for litera, den in gasite:
            asteptat = ddoc.ANEXE[litera]
            if not den.startswith(asteptat.split(" (")[0]):
                cade("13", f"{cfg['nume']}: Anexa {litera} se numește {den!r}, "
                           f"se aștepta {asteptat!r}")
        litere = [l for l, _ in gasite]
        if litere != sorted(litere):
            cade("13", f"{cfg['nume']}: anexele nu sunt în ordine: {litere}")
        for ext in (".docx", ".html"):
            frate = cale.replace(".md", ext)
            if not os.path.exists(frate):
                cade("13", f"{cfg['nume']}: lipsește {os.path.basename(frate)}")
    if not any(e.startswith("[13]") for e in esecuri):
        ok("13", f"cele {len(ddoc.DOCUMENTE)} documente au aceeași legendă, anexe "
                 f"canonice și cele trei formate (.md, .docx, .html)")


# ----------------------------------------------------------------------- poarta 16
def poarta_repartizare(wb):
    """Poarta 16 — o sursă care alimentează mai multe destinații nu pierde în cusătură.

    Poarta 12 verifică perechea document ↔ sursa lui. Când o singură sursă hrănește
    patru destinații, fiecare document poate trece poarta 12 separat în timp ce
    material cade ÎNTRE ele. Riscul nu e „nimeni n-a luat-o”, ci „am crezut că a
    luat-o celălalt”.

    a) fiecare subsecțiune a sursei are destinație declarată în `date/repartizare.py`;
    b) fiecare fragment ajunge în destinația DECLARATĂ — nu doar undeva. Verificarea
       pe reuniunea destinațiilor ar spune că textul există pe undeva, ceea ce e exact
       întrebarea greșită.
    """
    foi = {ws.title: {conservare._normalizeaza(l)
                      for row in ws.iter_rows() for c in row
                      if isinstance(c.value, str) for l in c.value.split("\n")
                      if conservare._normalizeaza(l)}
           for ws in wb.worksheets}

    orfane, lipsa, fara_artefact = brep.verifica(foi)

    for t in orfane:
        cade("16", f"subsecțiune fără destinație declarată: {t[:80]}")
    if not orfane:
        ok("16", f"toate cele {len(drep.REPARTIZARE)} subsecțiuni ale sursei au "
                 f"destinație declarată")

    for dest in fara_artefact:
        cade("16", f"destinația {dest} nu are încă artefact — "
                   f"{drep.DESTINATII[dest][0]} lipsește")
    for titlu, dest, frag in lipsa[:8]:
        cade("16", f"{titlu[:44]} → {dest}: nu s-a regăsit acolo — {frag[:70]}")
    if lipsa:
        cade("16", f"total {len(lipsa)} fragmente nu au ajuns în destinația declarată")
    elif not fara_artefact:
        ok("16", f"tot conținutul sursei a ajuns în destinația declarată "
                 f"({len(drep.DESTINATII)} destinații)")


# ----------------------------------------------------------------------- poarta 17
def poarta_inchideri(wb):
    """Poarta 17 — checklistul de închidere și monografiile spun același lucru.

    Foaia „Închideri periodice” afirmă că un cont trebuie să ajungă într-o anumită
    stare. Dacă nicio monografie nu demonstrează starea aia, checklistul cere ceva ce
    sistemul nu arată nicăieri. Invers, dacă un cont cu rol în flux se golește într-o
    monografie dar nu apare în checklist, disciplina lunară uită de el.

    Ambele sensuri, ca la porțile 4 (matrice ↔ fluxuri) și 10 (catalog ↔ monografii).
    Golul se declară — în `GOLURI` pentru primul sens, în `FARA_CADENTA` pentru al
    doilea — dar nu poate rămâne tăcut.
    """
    if binch.NUME not in wb.sheetnames:
        cade("17", f"lipsește foaia „{binch.NUME}” — rulează `make plan`")
        return

    harta = binch.asertiuni(wb)

    # sensul 1: cont urmărit → există aserțiune, sau golul e declarat
    nedeclarate = []
    for cont, _, _ in dinch.CADENTA:
        for s_ in binch.simboluri(cont):
            if s_ not in harta and s_ not in dinch.GOLURI:
                nedeclarate.append((cont, s_))
    for cont, s_ in nedeclarate:
        cade("17", f"{s_} (din „{cont[:30]}”) nu e starea terminală a niciunui flux "
                   f"și nu e declarat în GOLURI")
    if not nedeclarate:
        ok("17", f"toate cele {len(dinch.CADENTA)} rânduri de cadență sunt ancorate în "
                 f"fluxuri sau au golul declarat ({len(dinch.GOLURI)} goluri)")

    # sensul 2: cont cu rol în flux care se golește → are cadență, sau scutire declarată
    ws = wb["Plan de conturi"]
    rol = {}
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(row=r, column=1).value or "").strip()
        if re.fullmatch(r"\d{3,4}([./]\d+)?", v):
            rol[v] = str(ws.cell(row=r, column=4).value or "").strip().lower()

    urmarite = set()
    for cont, _, _ in dinch.CADENTA:
        urmarite.update(binch.simboluri(cont))
    urmarite |= set(dinch.GOLURI)

    uitate = []
    for simbol, aparitii in harta.items():
        if simbol in urmarite or simbol in dinch.FARA_CADENTA:
            continue
        r = rol.get(simbol) or rol.get(simbol[:3]) or ""
        if "rol in flux" not in r and "rol în flux" not in r:
            continue                       # patrimonialele au voie să poarte sold
        if not any(re.search(rf"[Ss]old(ul)?\s+{re.escape(simbol)}[^=]{{0,14}}=\s*0", t)
                   for _, _, t in aparitii):
            continue                       # se golește? dacă nu, nu e treaba listei
        uitate.append((simbol, aparitii[0][0]))
    for simbol, fid in uitate:
        cade("17", f"{simbol} are rol în flux și se golește în {fid}, dar nu are "
                   f"cadență și nu e scutit în FARA_CADENTA")
    if not uitate:
        ok("17", f"niciun cont cu rol în flux care se golește nu a rămas fără cadență "
                 f"({len(dinch.FARA_CADENTA)} scutiri declarate)")


# ----------------------------------------------------------------------- poarta 18
def poarta_monografii():
    """Poarta 18 — cifrele din monografiile scrise în proză.

    Poarta 1 se uită doar în `date/`. Documentele conțin peste o sută de blocuri de
    monografie prin care au trecut deja două erori, prinse de citire, nu de o poartă.

    a) articolele COMPUSE (`%`) se echilibrează: liniile de continuare trebuie să
       însumeze totalul declarat pe rândul de cap. Articolul simplu, scris pe o linie,
       nu are ce dezechilibra — are o singură sumă, aceeași pe ambele părți — deci
       verificarea lui ar fi vacuă, iar raportarea lui ca „verificat” ar minți;
    b) aritmetica afirmată în text se verifică. Asta e ramura care ar fi prins eroarea
       rezervei legale din trainingul 2: articolul se echilibra (125 = 125), dar „5% din
       250” nu face 125. Echilibrul singur nu vede așa ceva.

    Blocurile care citează deliberat o eroare din notițele brute sunt scutite, declarate
    în `date/monografii.py`.
    """
    articole = dezechilibrate = scutite = simple = 0
    for cfg in ddoc.DOCUMENTE:
        cale = os.path.join(RADACINA, cfg["iesire"])
        if not os.path.exists(cale):
            cade("18", f"{cfg['nume']}: lipsește {cfg['iesire']}")
            continue
        for bloc in bmono.citeste(cale):
            corp = "\n".join(bloc["linii"])
            if any(frag in corp for frag in dmono.CITEAZA_EROARE):
                scutite += 1
                continue
            for a in bloc["articole"]:
                if not any(x is not None for _, x in a["debit"] + a["credit"]):
                    continue
                simple += not a["compus"]
                if not a["compus"]:
                    continue        # o singură sumă: echilibrat prin construcție
                articole += 1
                sd = sum(x for _, x in a["debit"] if x is not None)
                sc = sum(x for _, x in a["credit"] if x is not None)
                if abs(sd - sc) > 0.005:
                    dezechilibrate += 1
                    cade("18", f"{cfg['nume']}:{a['rand']} ΣD={sd:,.2f} ≠ ΣC={sc:,.2f} "
                               f"— {a['brut'][0].strip()[:60]}")
    if not dezechilibrate:
        ok("18", f"toate cele {articole} articole compuse din monografiile în proză se "
                 f"echilibrează; {simple} articole simple n-au ce dezechilibra "
                 f"({scutite} blocuri scutite: citează o eroare)")

    gresite = total = 0
    for cfg in ddoc.DOCUMENTE:
        cale = os.path.join(RADACINA, cfg["iesire"])
        if not os.path.exists(cale):
            continue
        for rand, text, calculat, scris in bmono.afirmatii(cale):
            total += 1
            prag = max(dmono.TOLERANTA_ABSOLUTA, abs(scris) * dmono.TOLERANTA_RELATIVA)
            if abs(calculat - scris) > prag:
                gresite += 1
                cade("18", f"{cfg['nume']}:{rand} aritmetică falsă — „{text}” dă "
                           f"{calculat:,.2f}, nu {scris:,.2f}")
    if not gresite:
        ok("18", f"toate cele {total} afirmații aritmetice din text se verifică")


# ----------------------------------------------------------------------- poarta 19
def poarta_marcaje():
    """Poarta 19 — legenda aplicată, nu doar declarată.

    Poarta 13 verifică FORMA legendei: că toate cele patru marcaje sunt definite, la fel
    în toate documentele. Nu verifică nimic despre folosirea lor — gol notat în
    documentul de parcurs de la bun început.

    ❓ e singurul marcaj cu înțeles testabil: „rămas deschis, vezi Anexa D”. Deci:

    a) un document pe care cel puțin o întrebare din `date/intrebari.py` îl privește
       trebuie să poarte ❓ în corp. Fără el, cititorul vede un document care pare
       tranșat, deși sistemul știe că nu e;
    b) un document care poartă ❓ trebuie să aibă întrebări deschise. Un marcaj care nu
       trimite nicăieri e mai rău decât lipsa lui: promite o anexă care nu-l explică.

    Legătura întrebare → document se DEDUCE din data din `sursa`; doar sursa din 19.08,
    care alimentează patru documente, are cheile scrise explicit.

    Rândul de legendă (`| ❓ | Rămas deschis … |`) nu se numără: e definiția marcajului,
    nu o folosire a lui.
    """
    # Doar întrebările DESCHISE cer marcaj: una la care s-a găsit răspuns nu mai e
    # provizorie, iar ❓-ul ei a devenit ✅ în proză.
    cu_intrebari = {}
    for _, q in dintr.toate():
        if q["raspuns"]:
            continue
        d = dintr.documentul(q)
        if d:
            cu_intrebari.setdefault(d, []).append(q)

    poarta_marcaj = {}
    for cfg in ddoc.DOCUMENTE:
        cale = os.path.join(RADACINA, cfg["iesire"])
        if not os.path.exists(cale):
            cade("19", f"{cfg['nume']}: lipsește {cfg['iesire']}")
            continue
        n = 0
        with open(cale, encoding="utf-8") as f:
            for linie in f:
                if "❓" not in linie or linie.strip().startswith("| ❓ |"):
                    continue
                n += 1
        poarta_marcaj[cfg["cheie"]] = n

    for cheie, intrebari in sorted(cu_intrebari.items()):
        if poarta_marcaj.get(cheie, 0) == 0:
            cade("19", f"{cheie} are {len(intrebari)} întrebări deschise, dar niciun "
                       f"marcaj ❓ în corp — documentul pare tranșat, deși nu e")
    for cheie, n in sorted(poarta_marcaj.items()):
        if n and cheie not in cu_intrebari:
            cade("19", f"{cheie} poartă {n} marcaje ❓ dar nu are nicio întrebare "
                       f"deschisă — marcajul nu trimite nicăieri")

    if not any(e.startswith("[19]") for e in esecuri):
        ok("19", f"marcajul ❓ e aplicat unde trebuie: {len(cu_intrebari)} documente cu "
                 f"întrebări deschise, {sum(poarta_marcaj.values())} marcaje în corp")


# ------------------------------------------------------------------- porțile 20-22
def poarta_conturi_in_plan(wb):
    """Poarta 20 — contul folosit într-un pas de flux există în plan.

    Un cont care apare într-o monografie dar lipsește din „Plan de conturi” rupe
    navigarea cont → flux exact pentru contul acela: pleci din plan și nu-l găsești.

    Două capcane, ambele întâlnite la măsurare:

    - foaia are ANTETE REPETATE pe secțiuni, iar coloanele diferă între catalogul de
      fluxuri și monografii. Prima măsurătoare a citit coloana „Sumă” drept „Cont C” și
      a raportat 49 de conturi lipsă, toate false — erau sumele;
    - analiticele se reduc la sintetic: `371.AM.21` se verifică drept `371`, pentru că
      planul ține sinteticul, nu fiecare analitic imaginabil.
    """
    plan = set()
    for r in wb["Plan de conturi"].iter_rows(min_col=1, max_col=1, values_only=True):
        v = str(r[0] or "").strip()
        if re.fullmatch(r"\d{3,4}([./]\d+)?", v):
            plan.add(v)

    re_cont = re.compile(r"\b(\d{3,4})(?:\.[A-Za-z0-9ĂÂÎȘȚăâîșț]+)*\b")
    folosite, cap = {}, None
    for r in wb["Fluxuri"].iter_rows(min_col=1, max_col=8, values_only=True):
        v = [str(x or "") for x in r]
        if v[0].strip() == "Flux ID":
            cap = v
            continue
        fid = v[0].strip()
        if not re.fullmatch(r"F-\d{3}", fid) or not cap or "Cont D" not in cap:
            continue
        for col in (cap.index("Cont D"), cap.index("Cont C")):
            for m in re_cont.finditer(v[col]):
                folosite.setdefault(m.group(1), set()).add(fid)

    lipsa = {c: f for c, f in folosite.items() if c not in plan and c[:3] not in plan}
    for cont, fl in sorted(lipsa.items()):
        cade("20", f"contul {cont} apare în {', '.join(sorted(fl)[:4])} dar lipsește "
                   f"din „Plan de conturi”")
    if not lipsa:
        ok("20", f"toate cele {len(folosite)} conturi din pașii fluxurilor există în plan")


def poarta_module_declara(wb_plan):
    """Poarta 21 — fluxurile declarate de module există cu adevărat.

    Poarta 6 verifică sensul invers: fiecare token `MOD_*` referit există în catalog.
    Golul e aici: `build/ancore.py` derivă TOATE ancorele din `CATALOG['fluxuri']`, deci
    un ID greșit acolo nu produce o eroare — produce o ancoră lipsă, tăcut.
    """
    fluxuri = set()
    for r in wb_plan["Fluxuri"].iter_rows(min_col=1, max_col=1, values_only=True):
        m = re.match(r"^(F-\d{3})\b", str(r[0] or "").strip())
        if m:
            fluxuri.add(m.group(1))

    rele = []
    for m in dmod.MODULE:
        for fid in re.findall(r"F-\d{2,3}", m.CATALOG.get("fluxuri", "")):
            nou = O.HARTA.get(fid, fid)
            if nou not in fluxuri:
                rele.append((m.COD, fid, nou))
    for cod, vechi, nou in rele:
        cade("21", f"{cod} declară fluxul {vechi}"
                   f"{'' if nou == vechi else f' (→ {nou})'}, care nu există")
    if not rele:
        ok("21", f"toate fluxurile declarate de cele {len(dmod.MODULE)} module există")


def poarta_readme():
    """Poarta 22 — blocul de cifre din README e cel pe care generatorul l-ar produce.

    `make tot` îl reface oricum. Poarta prinde cazul în care README-ul a fost comis fără
    build — adică exact cazul în care a și rămas în urmă: afirma 23 corelații când erau
    29, și 58 de conturi Tier A când 87 sunt clasificate și 39 detaliate.
    """
    with open(breadme.CALE, encoding="utf-8") as f:
        text = f.read()
    if breadme.START not in text or breadme.STOP not in text:
        cade("22", "README.md nu are marcajele blocului generat")
        return
    i, j = text.index(breadme.START), text.index(breadme.STOP) + len(breadme.STOP)
    pe_disc, asteptat = text[i:j], breadme.bloc()
    if pe_disc == asteptat:
        ok("22", "blocul de cifre din README e la zi")
        return
    a = [l for l in pe_disc.split("\n") if l.strip()]
    b = [l for l in asteptat.split("\n") if l.strip()]
    for x, y in zip(a, b):
        if x != y:
            cade("22", f"README, blocul de cifre: „{x[:70]}” ar trebui să fie „{y[:70]}”")
            break
    else:
        cade("22", f"README, blocul de cifre: {len(a)} rânduri pe disc vs. {len(b)} "
                   f"generate — rulează `make readme`")


def poarta_cifre_in_proza():
    """Poarta 24 — frazele cu cifră din workbook-uri spun cifra reală.

    Legenda afirma „16 module declarative” (erau 17), „cele 21 de întrebări rămase
    deschise” (erau 20), „81 conturi de serviciu” (80) și „Tier A (~55)” (87 clasificate,
    39 detaliate). Foaia Fluxuri își anunța catalogul cu „LISTA FLUXURILOR (~38)”, la
    trei rânduri deasupra celor 68 pe care le lista.

    Ce verifică: fiecare apariție a unui tipar din `cifre.FRAZE`. Ce NU verifică: un
    număr scris în proză pentru care nu există tipar acolo. Acoperirea e exact acea
    listă — un `\d+ conturi` generic ar fi picat pe „Adăugate … 21 de conturi Tier A”,
    care e o afirmație istorică despre o adăugire, nu un total.
    """
    gresite = bcifre.fraze_gresite([PLAN, MODULE])
    if not gresite:
        ok("24", f"toate frazele cu cifră din cele două workbook-uri "
                 f"({len(bcifre.FRAZE)} tipare) spun cifra reală")
        return
    for fis, foaie, cel, fraza, gasit, astept in gresite[:6]:
        cade("24", f"{fis} · {foaie}!{cel}: „{fraza}” — sunt {astept}, nu {gasit}")


def poarta_foi_de_modul(wb_plan):
    """Porțile 23 și 25 — foile de modul: numite corect, și toate generate.

    **23** e împotriva clasei de erori „Declarații_TVA”: `Index module` numea, pentru
    cele șapte module de sămânță, foi care nu existau (`Declarații_APROV`,
    `Declarații_INTER`, `Declarații_EX`, `Declarații_NEUT`) și omitea `Reguli_`, pe care
    modulele alea chiar le aveau — o instrucțiune de navigare greșită în foaia al cărei
    singur rost e navigarea. Verificarea merge în ambele sensuri, pentru că doar una
    n-ar fi de-ajuns: o foaie fără intrare de catalog e la fel de invizibilă ca o
    intrare care trimite în gol.

    **25** e poarta care ține câștigul portării. Fără ea, o sămânță viitoare ar putea
    reintroduce foi întreținute de mână și nimeni n-ar observa până când n-ar diverge.
    """
    wbm = openpyxl.load_workbook(MODULE, read_only=True)
    reale = set(wbm.sheetnames)
    wbm.close()

    # ---- 23a: fiecare foaie pe care catalogul o numește există cu adevărat
    lipsa = []
    for m in dmod.MODULE:
        for nume in dmcomun.foi(m):
            if nume not in reale:
                lipsa.append((m.COD, nume))
    # …iar coloana din `Index module` spune exact aceleași nume
    coloana = {}
    for r in wb_plan["Index module"].iter_rows(values_only=True):
        cod = str(r[0] or "").strip()
        if cod.startswith("MOD_") and len(r) > 4:
            coloana[cod] = [x.strip() for x in str(r[4] or "").split(",") if x.strip()]
    for m in dmod.MODULE:
        scrise, asteptate = coloana.get(m.COD, []), dmcomun.foi(m)
        if scrise != asteptate:
            lipsa.append((m.COD, f"Index module scrie {scrise!r}, nu {asteptate!r}"))

    # ---- 23b: fiecare foaie de modul are exact o intrare de catalog
    declarate = {n for m in dmod.MODULE for n in dmcomun.foi(m)}
    orfane = sorted(s for s in reale
                    if any(s.startswith(p + "_") for p in
                           dmcomun.PREFIXE_STANDARD + ("Verificări", "Abateri",
                                                       "Registru"))
                    and s not in declarate)

    if lipsa or orfane:
        for cod, ce in lipsa[:5]:
            cade("23", f"{cod}: foaia declarată nu există — {ce}")
        for s_ in orfane[:5]:
            cade("23", f"foaia {s_!r} nu are nicio intrare de catalog")
    else:
        ok("23", f"toate cele {len(declarate)} foi de modul sunt numite corect în "
                 f"catalog, în ambele sensuri")

    # ---- 25: nicio foaie de modul nemaigenerată din `date/`
    negenerate = sorted(reale - declarate - {"Instructiuni", "Parametri",
                                             "CatalogModule"})
    if negenerate:
        for s_ in negenerate[:5]:
            cade("25", f"foaia {s_!r} vine din sămânță — n-are generator în date/module")
    else:
        ok("25", f"toate cele {len(reale)} foi ale workbook-ului de module se generează "
                 f"din date/ — nicio foaie nu mai vine din sămânță")


def poarta_diacritice(wb):
    """Poarta 26 — harta de diacritice e aplicată complet.

    Prinde harta aplicată pe jumătate: dacă un cuvânt din partea STÂNGĂ a hărții mai
    apare în foile acoperite, ori aplicarea a fost sărită, ori o etapă de mai târziu a
    rescris celula peste ea.

    Verifică și partea care contează cel mai mult: niciun SIMBOL de cont n-a fost atins.
    Cele două foi țin 257 de conturi, iar un simbol schimbat ar rupe tăcut navigarea
    cont → flux și poarta 20.
    """
    gasite = []
    for nume in ddiac.FOI:
        ws = wb[nume]
        col_simbol = {c.column for row in ws.iter_rows() for c in row
                      if isinstance(c.value, str)
                      and c.value.strip().lower() in ddiac.CAP_SIMBOL}
        for row in ws.iter_rows():
            for c in row:
                if not isinstance(c.value, str) or c.column in col_simbol:
                    continue
                if bdiac.RE_CONT.match(c.value.strip()):
                    continue
                for w in bdiac._RE_CUVANT.findall(c.value):
                    if w.lower() in ddiac.HARTA:
                        gasite.append((nume, c.coordinate, w))

    simboluri = sum(1 for nume in ddiac.FOI for row in wb[nume].iter_rows()
                    for c in row[:1]
                    if isinstance(c.value, str)
                    and bdiac.RE_CONT.match(c.value.strip()))

    if gasite:
        for nume, coord, w in gasite[:5]:
            cade("26", f"{nume}!{coord}: „{w}” e în hartă, dar n-a fost înlocuit")
        return
    ok("26", f"harta de diacritice ({len(ddiac.HARTA)} intrări, "
             f"{len(ddiac.EXCEPTII)} omografe declarate) e aplicată complet pe cele "
             f"{len(ddiac.FOI)} foi; {simboluri} simboluri de cont neatinse")


def main():
    if not os.path.exists(PLAN):
        raise SystemExit("Rulează întâi `make build` — lipsește dist/Plan_de_conturi_...xlsx")
    poarta_fluxuri()
    poarta_factori()
    wb = openpyxl.load_workbook(PLAN)
    poarta_acoperire(wb)
    poarta_corelatii()
    poarta_module()
    poarta_conservare()
    poarta_catalog(wb)
    poarta_nume_definite()
    poarta_structura(wb)
    poarta_documente()
    poarta_parcurs(wb)
    poarta_repartizare(wb)
    poarta_inchideri(wb)
    poarta_monografii()
    poarta_marcaje()
    poarta_conturi_in_plan(wb)
    poarta_module_declara(wb)
    poarta_readme()
    poarta_cifre_in_proza()
    poarta_foi_de_modul(wb)
    poarta_diacritice(wb)

    print("\n".join(note))
    if esecuri:
        print(f"\n✘ {len(esecuri)} eșecuri:")
        for e in esecuri:
            print("   " + e)
        sys.exit(1)
    print(f"\n✔ toate porțile trecute ({len(FLUXURI)} fluxuri, {len(CORELATII)} corelații, "
          f"{len(ANALITICE)} analitice)")


if __name__ == "__main__":
    main()

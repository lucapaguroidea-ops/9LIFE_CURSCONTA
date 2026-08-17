"""Mutarea rândurilor între foi, cu stiluri și îmbinări intacte.

Reordonarea unei foi înseamnă rescrierea ei într-o foaie nouă, în altă ordine.
`openpyxl` nu are așa ceva: `insert_rows` mută valorile dar lasă îmbinările pe loc,
iar scrierile în afara colțului stânga-sus al unei îmbinări se pierd tăcut la salvare.

Modulul ăsta lucrează la nivel de RÂND capturat: valoare + stil + înălțime + span-ul
de îmbinare. Un rând capturat se poate scrie oriunde, de câte ori vrei.
"""
from copy import copy


class Rand:
    """Un rând capturat: valorile, stilurile, înălțimea și îmbinarea lui."""

    __slots__ = ("celule", "inaltime", "span", "stil_nou")

    def __init__(self, celule, inaltime=None, span=None, stil_nou=None):
        self.celule = celule          # listă de (coloană, valoare, stil)
        self.inaltime = inaltime
        self.span = span              # (col_start, col_end) dacă rândul era îmbinat
        self.stil_nou = stil_nou      # {font, fill, align} pentru rânduri create acum

    def text(self, col=1):
        for c, v, _ in self.celule:
            if c == col:
                return str(v) if v is not None else ""
        return ""

    def valori(self):
        return {c: v for c, v, _ in self.celule}

    def gol(self):
        return not any(str(v).strip() for _, v, _ in self.celule if v is not None)

    def cu_valoare(self, col, valoare):
        """Copie a rândului, cu o singură celulă schimbată (stilul se păstrează)."""
        gasit = False
        noi = []
        for c, v, s in self.celule:
            if c == col:
                noi.append((c, valoare, s))
                gasit = True
            else:
                noi.append((c, v, s))
        if not gasit:
            stil = self.celule[0][2] if self.celule else None
            noi.append((col, valoare, stil))
        return Rand(noi, self.inaltime, self.span)

    def transformat(self, fn):
        """Copie a rândului cu fn() aplicată pe fiecare valoare de tip text."""
        return Rand([(c, fn(v) if isinstance(v, str) else v, s) for c, v, s in self.celule],
                    self.inaltime, self.span)


def citeste(ws):
    """Toată foaia, ca listă de Rand (indexată de la 0 = rândul 1)."""
    imbinari = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row == rng.max_row:
            imbinari[rng.min_row] = (rng.min_col, rng.max_col)

    out = []
    for r in range(1, ws.max_row + 1):
        celule = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None or cell.has_style:
                celule.append((c, cell.value, copy(cell._style)))
        dim = ws.row_dimensions.get(r)
        out.append(Rand(celule, dim.height if dim else None, imbinari.get(r)))
    return out


def scrie(ws, randuri, *, start=1):
    """Scrie rândurile capturate începând de la `start`. Întoarce primul rând liber."""
    from build import stil as S
    r = start
    for rand in randuri:
        for col, valoare, stil in rand.celule:
            if rand.stil_nou is not None:
                S.scrie(ws, r, col, valoare, **rand.stil_nou)
                continue
            cell = ws.cell(row=r, column=col)
            cell.value = valoare
            if stil is not None:
                cell._style = copy(stil)
        if rand.inaltime is not None:
            ws.row_dimensions[r].height = rand.inaltime
        if rand.span:
            c1, c2 = rand.span
            if c2 > c1:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        r += 1
    return r


def inlocuieste_foaia(wb, nume, randuri):
    """Rescrie o foaie de la zero cu rândurile date, păstrându-i poziția și setările.

    Lățimile de coloană, panourile înghețate și autofiltrul se transferă. Foaia veche
    se șterge abia după ce cea nouă e completă, iar la final se redenumește — ca orice
    referință prin nume din alte foi să rămână validă.
    """
    vechea = wb[nume]
    index = wb.sheetnames.index(nume)
    latimi = {k: (v.width, v.hidden) for k, v in vechea.column_dimensions.items()}
    freeze = vechea.freeze_panes
    avea_filtru = vechea.auto_filter.ref is not None
    latime_max = max((c for rand in randuri for c, _, _ in rand.celule), default=1)

    temp = wb.create_sheet(f"__{nume}__", index + 1)
    scrie(temp, randuri)
    for k, (w, hidden) in latimi.items():
        temp.column_dimensions[k].width = w
        temp.column_dimensions[k].hidden = hidden
    if freeze:
        temp.freeze_panes = freeze

    del wb[nume]
    temp.title = nume
    wb.move_sheet(nume, offset=index - wb.sheetnames.index(nume))
    if avea_filtru:
        from openpyxl.utils import get_column_letter
        ultim = len(randuri)
        temp.auto_filter.ref = f"A1:{get_column_letter(latime_max)}{ultim}"
    return temp

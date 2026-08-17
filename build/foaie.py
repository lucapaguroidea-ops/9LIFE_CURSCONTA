"""Constructor de foaie care ține evidența rândurilor și întoarce referințe.

Motivul: modulele declarative sunt pline de referințe încrucișate
(`=Declarații_X!B22`, `=C8+C14+C21`). Scrise de mână, se rup la prima linie
adăugată. Aici fiecare metodă întoarce referința celulei pe care tocmai a
scris-o, așa că formulele se compun din variabile, nu din coordonate literale.
"""
from openpyxl.utils import get_column_letter

from build import stil


class Foaie:
    def __init__(self, wb, nume, latimi=None):
        if nume in wb.sheetnames:
            raise SystemExit(f"Foaia {nume!r} există deja — cod de modul duplicat?")
        self.ws = wb.create_sheet(nume)
        self.nume = nume
        self.r = 1
        if latimi:
            for col, w in latimi.items():
                self.ws.column_dimensions[col].width = w

    # ------------------------------------------------------------------ intern
    def _scrie(self, col, val, font=None, fill=None, align=stil.A_WRAP_TOP):
        stil.scrie(self.ws, self.r, col, val, font=font, fill=fill, align=align)
        return f"{get_column_letter(col)}{self.r}"

    def ref(self, coord):
        """Referință absolută către o celulă din ACEASTĂ foaie, utilizabilă din alta.

        Diacriticele nu cer ghilimele (fișierul original folosește
        `Declarații_INCHIDERE_TVA!B8` neîncadrat); spațiul și cratima, da.
        """
        nume = self.nume
        if any(ch in nume for ch in " -()"):
            nume = f"'{nume}'"
        return f"{nume}!{coord}"

    # ------------------------------------------------------------------ scriere
    def titlu(self, text):
        self._scrie(1, text, font=stil.F_TITLU)
        self.r += 1

    def nota(self, text):
        self._scrie(1, text, font=stil.F_NOTA)
        self.r += 1

    def avert(self, text):
        self._scrie(1, text, font=stil.F_AVERTISMENT)
        self.r += 1

    def sectiune(self, text):
        self._scrie(1, text, font=stil.F_TITLU_BLOC)
        self.r += 1

    def gol(self, n=1):
        self.r += n

    def kv(self, eticheta, valoare, tip="input", nota=None):
        """Linie etichetă / valoare. Întoarce referința celulei de valoare (col. B)."""
        self._scrie(1, eticheta, font=stil.F_NORMAL)
        if tip == "input":
            ref = self._scrie(2, valoare, font=stil.F_INPUT, fill=stil.FILL_INPUT)
        else:
            ref = self._scrie(2, valoare, font=stil.F_NORMAL)
        if nota:
            self._scrie(3, nota, font=stil.F_NOTA)
        self.r += 1
        return ref

    def cap(self, capete):
        for i, h in enumerate(capete, start=1):
            self._scrie(i, h, font=stil.F_CAP_TABEL_ALB, fill=stil.FILL_ANTET,
                        align=stil.A_CENTER)
        self.r += 1

    def rand(self, valori, fill=None, font=None):
        """Scrie un rând de tabel. Întoarce dict {litera_coloanei: referință}."""
        refs = {}
        for i, v in enumerate(valori, start=1):
            if v is None:
                continue
            litera = get_column_letter(i)
            refs[litera] = self._scrie(i, v, font=font or stil.F_NORMAL, fill=fill)
        self.r += 1
        return refs

    def check(self, eticheta, formula, formula_verdict):
        """Linie de control: etichetă · valoare · verdict OK/EROARE."""
        self._scrie(1, eticheta, font=stil.F_NORMAL)
        ref = self._scrie(2, formula, font=stil.F_NORMAL)
        self._scrie(3, formula_verdict, font=stil.F_NORMAL)
        self.r += 1
        return ref

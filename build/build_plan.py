"""Generează dist/Plan_de_conturi_ROL_Analitice_Fluxuri_SAGA.xlsx.

Strategie: EXTINDERE PE BAZĂ DE SEED. Se încarcă workbook-ul original din
`surse/training-4-2026-08-14/` și se adaugă conținutul nou din `date/`:

- `Plan de conturi`   — actualizări IN LOC pe rândurile existente (fără inserări de
                        rânduri, care ar strica autofiltrul), plus conturile lipsă
                        adăugate la final, plus corecția denumirii lui 235
- `Analitice (Tier A)`— rânduri noi la finalul tabelului
- `Fluxuri`           — o secțiune „TRANȘA 4” la final: catalog + blocuri de pași
- `Corelații de control` — secțiune C-13…C-22 la final
- `Matrice acoperire` — secțiune nouă + rezolvarea marcajelor PARȚIAL
- `Legendă`, `Index module` — note de etapă

Rulare:  python build/build_plan.py
"""
import os
import sys

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from build import stil  # noqa: E402
from build.recalc import recalc  # noqa: E402
from date import ANALITICE, CORELATII, FLUXURI, plan as dplan  # noqa: E402
from date import reformulari as dreform  # noqa: E402
from date import ordine as O  # noqa: E402
from date import module as dmodule  # noqa: E402
from build import ancore, foaie_intrebari, istoric, renumeroteaza  # noqa: E402
from build import reordoneaza, reordoneaza_foi  # noqa: E402
from date.comun import ro  # noqa: E402

RADACINA = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SEED = os.path.join(RADACINA, "surse", "training-4-2026-08-14",
                    "Plan_de_conturi_ROL_Analitice_Fluxuri_SAGA.xlsx")
IESIRE = os.path.join(RADACINA, "dist", "Plan_de_conturi_ROL_Analitice_Fluxuri_SAGA.xlsx")

# coloanele foii „Plan de conturi” (1-indexate)
COL_SIMBOL, COL_DENUMIRE, COL_FCT = 1, 2, 3
COL_NATURA, COL_SUBTIP, COL_OBS = 4, 5, 6
COL_ANALITICE, COL_FACTOR, COL_FLUX, COL_TIER = 7, 8, 9, 10


def _contopeste(vechi, nou):
    """Actualizează o celulă PĂSTRÂND ce era acolo.

    Regula implicită e contopirea, nu suprascrierea: dacă rândul avea deja o notă,
    ea rămâne și noul conținut se adaugă după ea. Suprascrierea e permisă doar
    pentru textele declarate în `date/reformulari.py`, cu motiv.

    Așa, clasa asta de pierdere devine imposibilă, nu doar reparată o dată.
    """
    v = str(vechi or "").strip()
    n = str(nou or "").strip()
    if not v or v == "—":
        return n
    if v in dreform.SUPRASCRIE:
        return dreform.SUPRASCRIE[v]
    if not n or n == v or v in n:
        return n or v
    return f"{v} — {n}"


def _primul_rand_liber(ws, col=1):
    """Primul rând complet gol de la finalul foii."""
    r = ws.max_row
    while r > 0 and all(ws.cell(row=r, column=c).value in (None, "")
                        for c in range(1, ws.max_column + 1)):
        r -= 1
    return r + 1


def _titlu_sectiune(ws, rand, text, latime):
    ws.merge_cells(start_row=rand, start_column=1, end_row=rand, end_column=latime)
    stil.scrie(ws, rand, 1, text, font=stil.F_TITLU_BLOC)
    return rand + 1


def _cap_tabel(ws, rand, capete, *, alb=False):
    for i, h in enumerate(capete, start=1):
        stil.scrie(ws, rand, i, h,
                   font=stil.F_CAP_TABEL_ALB if alb else stil.F_CAP_TABEL,
                   fill=stil.FILL_ANTET if alb else None,
                   align=stil.A_CENTER if alb else None)
    return rand + 1


def _nota(ws, rand, text, latime):
    ws.merge_cells(start_row=rand, start_column=1, end_row=rand, end_column=latime)
    stil.scrie(ws, rand, 1, text, font=stil.F_NOTA)
    return rand + 1


# ---------------------------------------------------------------- Plan de conturi
def extinde_plan_conturi(wb):
    ws = wb["Plan de conturi"]
    indice = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=COL_SIMBOL).value
        if v is not None:
            indice[str(v).strip()] = r

    # 1. corecții de denumire pe rânduri existente
    for c in dplan.CORECTII:
        r = indice.get(c["simbol"])
        if r is None:
            raise SystemExit(f"Corecție imposibilă: contul {c['simbol']} nu există în plan")
        actual = str(ws.cell(row=r, column=COL_DENUMIRE).value or "").strip()
        if actual != c["vechi"]:
            raise SystemExit(
                f"Corecție refuzată pentru {c['simbol']}: denumirea din fișier este "
                f"{actual!r}, nu {c['vechi']!r}. Verifică manual înainte de a suprascrie."
            )
        ws.cell(row=r, column=COL_DENUMIRE, value=c["nou"])
        obs = str(ws.cell(row=r, column=COL_OBS).value or "").strip()
        marca = f"⚠ Corectat 2026: {c['motiv']}"
        ws.cell(row=r, column=COL_OBS, value=(obs + " " + marca).strip())

    # 2. actualizări în loc — prin CONTOPIRE, nu prin suprascriere
    lipsa = []
    for simbol, camp in dplan.ACTUALIZARI.items():
        r = indice.get(simbol)
        if r is None:
            lipsa.append(simbol)
            continue
        for cheie, col in (("analitice", COL_ANALITICE), ("factor", COL_FACTOR),
                           ("flux", COL_FLUX), ("tier", COL_TIER)):
            if cheie in camp:
                ws.cell(row=r, column=col, value=_contopeste(
                    ws.cell(row=r, column=col).value, camp[cheie]))
    if lipsa:
        raise SystemExit(f"Conturi din ACTUALIZARI care nu există în plan: {lipsa}")

    # 3. conturi noi, la finalul tabelului
    rand = _primul_rand_liber(ws)
    rand += 1
    rand = _titlu_sectiune(ws, rand, "CONTURI ADĂUGATE ÎN ETAPA 4 (lipseau din planul inițial)", 10)
    for c in dplan.CONTURI_NOI:
        for col, val in (
            (COL_SIMBOL, c["simbol"]), (COL_DENUMIRE, c["denumire"]), (COL_FCT, c["fct"]),
            (COL_NATURA, c["natura"]), (COL_SUBTIP, c["subtip"]), (COL_OBS, c["observatie"]),
            (COL_ANALITICE, c["analitice"]), (COL_FACTOR, c["factor"]),
            (COL_FLUX, c["flux"]), (COL_TIER, c["tier"]),
        ):
            stil.scrie(ws, rand, col, val, font=stil.F_NORMAL, align=stil.A_WRAP)
        rand += 1
    return ws


# --------------------------------------------------------------- Analitice Tier A
def extinde_analitice(wb):
    ws = wb["Analitice (Tier A)"]
    rand = _primul_rand_liber(ws) + 1
    rand = _titlu_sectiune(ws, rand, "TRANȘA 4 — capitaluri (training 07.08.2026) și "
                                     "imobilizări (training 12.08.2026)", 7)
    for a in ANALITICE:
        for col, val in enumerate(
            [a["simbol"], a["denumire"], a["structura"], a["factor"],
             a["rupe"], a["nota"], a["flux"]], start=1
        ):
            stil.scrie(ws, rand, col, val, font=stil.F_NORMAL, align=stil.A_WRAP)
        rand += 1


# ------------------------------------------------------------------------ Fluxuri
def _randeaza_sume(p):
    """Întoarce (cont_d, cont_c, suma) în convenția foii originale."""
    dr, cr = p["dr"], p["cr"]
    if not dr and not cr:
        return "—", "—", "—"
    nd, nc = len(dr), len(cr)
    cont_d = "\n".join(c for c, _ in dr)
    cont_c = "\n".join(c for c, _ in cr)
    if nd > 1 and nc == 1:
        suma = "\n".join(ro(s) for _, s in dr)
    elif nc > 1 and nd == 1:
        suma = "\n".join(ro(s) for _, s in cr)
    elif nd == 1 and nc == 1:
        suma = ro(dr[0][1])
    else:
        suma = "\n".join(ro(s) for _, s in dr)
    return cont_d, cont_c, suma


def extinde_fluxuri(wb):
    ws = wb["Fluxuri"]
    rand = _primul_rand_liber(ws) + 2

    rand = _titlu_sectiune(
        ws, rand,
        "TRANȘA 4 — CAPITALURI ȘI IMOBILIZĂRI (F-45…F-62) — Etapa 4", 9)
    rand = _nota(ws, rand,
                 "Din trainingurile 07.08.2026 (capitaluri, credite, leasing, provizioane) și "
                 "12.08.2026 (imobilizări). Cifrele sunt cele din documentele revizuite, nu din "
                 "notițele brute. ★ = flux didactic, cu pas revelator identificat.", 9)
    rand += 1

    # --- catalogul tranșei 4
    rand = _cap_tabel(ws, rand, ["Flux ID", "Tranșă", "Denumire flux", "Didactic",
                                 "Rol(uri) revelate", "Conturi cheie", "Status",
                                 "Note / Declarativ legat"], alb=True)
    for f in FLUXURI:
        fill = stil.FILL_DIDACTIC if f["didactic"].startswith("★") else stil.FILL_NEDIDACTIC
        for col, val in enumerate([f["id"], f["transa"], f["denumire"], f["didactic"],
                                   f["roluri"], f["conturi"], f["status"], f["note"]], start=1):
            stil.scrie(ws, rand, col, val, font=stil.F_NORMAL, fill=fill, align=stil.A_WRAP)
        rand += 1
    rand += 2

    # --- blocurile de pași
    for f in FLUXURI:
        rand = _titlu_sectiune(ws, rand, f["titlu"], 9)
        rand = _cap_tabel(ws, rand, ["Flux ID", "Pas", "Document", "Explicație", "Cont D",
                                     "Cont C", "Sumă (lei)", "Rol revelat", "Declarativ"])
        for p in f["pasi"]:
            cont_d, cont_c, suma = _randeaza_sume(p)
            fill = stil.FILL_PAS_REVELATOR if p["revelator"] else stil.FILL_PAS
            for col, val in enumerate([f["id"], p["nr"], p["doc"], p["explicatie"],
                                       cont_d, cont_c, suma, p["rol"], p["declarativ"]], start=1):
                stil.scrie(ws, rand, col, val, font=stil.F_NORMAL, fill=fill, align=stil.A_WRAP)
            rand += 1
        if f["principiu"]:
            rand = _nota(ws, rand, "Principiul: " + f["principiu"], 9)
        rand += 1


# ------------------------------------------------------------- Corelații de control
def extinde_corelatii(wb):
    ws = wb["Corelații de control"]
    rand = _primul_rand_liber(ws) + 2
    rand = _titlu_sectiune(
        ws, rand, "TRANȘA 4 — CORELAȚII PE CAPITALURI ȘI IMOBILIZĂRI (C-13…C-22)", 7)
    rand = _nota(ws, rand,
                 "Aceeași utilizare ca la C-01…C-12: dacă o corelație nu ține, cauți întâi în coloana "
                 "LEGITIM; dacă nimic de acolo nu explică ruptura, greșeala e în coloana SUSPECT.", 7)
    rand += 1
    rand = _cap_tabel(ws, rand, ["ID", "Corelație (formulă)", "Unde se verifică",
                                 "Ce o rupe LEGITIM", "Ce o rupe SUSPECT",
                                 "Flux / modul legat", "Severitate dacă e ruptă suspect"], alb=True)
    for c in CORELATII:
        for col, val in enumerate([c["id"], c["formula"], c["unde"], c["legitim"],
                                   c["suspect"], c["flux"], c["severitate"]], start=1):
            stil.scrie(ws, rand, col, val, font=stil.F_NORMAL, align=stil.A_WRAP_TOP)
        rand += 1


# --------------------------------------------------------------- Matrice acoperire
def extinde_matrice(wb):
    ws = wb["Matrice acoperire"]
    # rezolvă marcajele PARȚIAL acoperite acum de tranșa 4
    rezolvate = {
        "28x": ("F-26, F-31, F-59, F-60, F-61", "F-61 oglinda 21x↔28x; F-59/F-60 descărcare la ieșire"),
        "681/781": ("F-26, F-29, F-50, F-52, F-53, F-54, F-51", "F-50 plafon 1.500 lei/lună; F-51 reluare 7812"),
        "167": ("F-50", "F-50 pas 9: 167 = scadențar, 1:1 cu contractul"),
        "421/431/444/436": (
            "F-32, F-52, F-58",
            "Lanțul complet în MOD_SALARII, cu cifre verificate contra statului real "
            "din 31.07.2026; F-52/F-58 acoperă ipostaza de cost capitalizat"),
        "641/642/646": (
            "F-32, F-52, F-58",
            "MOD_SALARII acoperă 641 salarii, 642 tichete și 646 CAM; limitările "
            "rămase sunt declarate în Reguli_SALARII, tabelul C"),
    }
    gasite = set()
    for r in range(1, ws.max_row + 1):
        simbol = ws.cell(row=r, column=1).value
        if simbol and str(simbol).strip() in rezolvate:
            s = str(simbol).strip()
            fluxuri, nota = rezolvate[s]
            ws.cell(row=r, column=4, value=fluxuri)
            ws.cell(row=r, column=5, value=nota)
            ws.cell(row=r, column=6, value="NU")
            gasite.add(s)
    lipsa = set(rezolvate) - gasite
    if lipsa:
        raise SystemExit(f"Rânduri PARȚIAL de rezolvat, negăsite în matrice: {sorted(lipsa)}")

    rand = _primul_rand_liber(ws) + 2
    rand = _titlu_sectiune(ws, rand, "TRANȘA 4 — conturi de clasa 1 și 2 (F-45…F-62)", 6)
    rand = _cap_tabel(ws, rand, ["Simbol", "Denumire scurtă", "Tier", "Flux(uri)",
                                 "Pas revelator / note", "Gol?"], alb=True)
    for row in dplan.MATRICE:
        for col, val in enumerate(row, start=1):
            stil.scrie(ws, rand, col, val, font=stil.F_NORMAL, align=stil.A_WRAP)
        rand += 1
    rand += 1
    _nota(ws, rand,
          "Marcajele PARȚIAL de la 28x, 681/781 și 167 din etapele anterioare au trecut pe NU "
          "prin fluxurile F-50, F-59, F-60 și F-61.", 6)


# ------------------------------------------------------------------ Legendă / Index
def actualizeaza_legenda(wb):
    ws = wb["Legendă"]
    rand = _primul_rand_liber(ws) + 1
    linii = [
        ("ETAPA 4 — EXTINDERE LA TRAININGURILE 2 ȘI 3", stil.F_TITLU_BLOC),
        ("Sursă: notițele revizuite din 07.08.2026 (capitaluri) și 12.08.2026 (imobilizări). "
         "Adăugate 18 fluxuri (F-45…F-62), 10 corelații (C-13…C-22) și 21 de conturi Tier A din "
         "clasele 1 și 2. Total fluxuri: 62.", stil.F_NORMAL),
        ("Fluxuri didactice noi: F-46 (121 = 129, nu invers), F-47 (1174 ocolește 121), "
         "F-49 (cele două momente de curs), F-50 (leasing 50% — trei limitări pe baze diferite), "
         "F-52 (721, nu 711), F-54 (cod 27 în D394), F-58 (722, nu 711), F-59 (valoare rămasă vs. venit), "
         "F-61 (control analitic ↔ sintetic — singurul flux de procedură, fără note contabile).",
         stil.F_NORMAL),
        ("Corecție aplicată pe planul de conturi: contul 235 avea denumirea contului 233. "
         "Conform OMFP 1802/2014: 231 corporale în curs, 233 necorporale în curs, "
         "235 investiții imobiliare în curs. Contul 233 a fost adăugat.", stil.F_AVERTISMENT),
        ("Erori din notițele brute NEreintroduse (corectate deja în documentele revizuite): "
         "7815 nu există (corect 6811 = 2805/2808); taxele vamale se capitalizează, nu pe 635; "
         "softul intern prin 721; salariile din 231 prin 722; 2114 este 214; "
         "CASCO nedeductibil pe 613.NED, nu 615; typo-urile 1067 (=167) și 4424 (=4426) la leasing.",
         stil.F_NORMAL),
        ("Reproductibilitate: workbook-ul se generează cu `make build` din surse/ + date/. "
         "Porțile de calitate se verifică cu `make verifica`.", stil.F_NOTA),
    ]
    for text, font in linii:
        ws.merge_cells(start_row=rand, start_column=1, end_row=rand, end_column=7)
        stil.scrie(ws, rand, 1, text, font=font, align=stil.A_WRAP_TOP)
        rand += 1


def actualizeaza_index(wb):
    ws = wb["Index module"]
    rand = _primul_rand_liber(ws) + 1
    rand = _titlu_sectiune(ws, rand, "MODULE ADĂUGATE ÎN ETAPA 4", 6)
    rand = _cap_tabel(ws, rand, ["Cod modul", "Fluxuri acoperite", "Ce face", "Status",
                                 "Foile din modul", "Când îl rulezi"], alb=True)
    # derivat din date/module — altfel indexul rămâne în urma modulelor reale, exact
    # cum s-a întâmplat cu MOD_LEASING_FIN, care figura și „EXEMPLU EXTERN”, și
    # „IMPLEMENTAT”, în două rânduri care se contraziceau
    module = [
        (m.COD, m.CATALOG["fluxuri"], m.CATALOG["blocuri"], "IMPLEMENTAT",
         ", ".join(f"{f}_{m.COD.removeprefix('MOD_')}"
                   for f in ("Declarații", "Reguli", "Jurnale", "NotaExport")),
         m.CATALOG["tip"])
        for m in dmodule.MODULE
    ]
    # modulele care există deja în index se actualizează în loc — un modul are un
    # singur rând, altfel indexul se contrazice singur (MOD_LEASING_FIN apărea de
    # două ori: „EXEMPLU EXTERN” și „IMPLEMENTAT”)
    existente = {}
    for r in range(1, ws.max_row + 1):
        cod = str(ws.cell(row=r, column=1).value or "").strip()
        if cod.startswith("MOD_"):
            existente[cod] = r
    for m in module:
        r = existente.get(m[0], rand)
        for col, val in enumerate(m, start=1):
            stil.scrie(ws, r, col, val, font=stil.F_NORMAL, align=stil.A_WRAP)
        if m[0] not in existente:
            rand += 1
    rand += 1
    _nota(ws, rand,
          "MOD_CAPITALURI nu dublează MOD_INCHIDERE_EX: acela închide clasele 6 și 7 pe 121, "
          "acesta repartizează rezultatul obținut (129/1061/1171).", 6)


def ordoneaza_canonic(wb):
    """Faza finală: din ordinea adăugirii în ordinea planului de conturi.

    Ordinea contează. Istoricul se construiește ÎNAINTE de renumerotare, iar
    renumerotarea îl ocolește — altfel și-ar rescrie propriul tabel de echivalență.
    """
    orfane, resturi, absorbite, statusuri = reordoneaza.rescrie(wb)

    dupl = reordoneaza_foi.reordoneaza_tabel(
        wb, "Analitice (Tier A)", latime=7,
        nota="Ordonat pe simbol de cont, ca planul de conturi. Rândurile duplicate "
             "apărute la extinderi succesive au fost contopite, fără pierdere de text.")
    dupl += reordoneaza_foi.reordoneaza_tabel(
        wb, "Matrice acoperire", latime=6,
        nota="Ordonat pe simbol de cont. PARȚIAL rămâne doar unde acoperirea chiar "
             "este parțială — vezi GOLURI_ACCEPTATE în date/plan.py.")

    mutate = reordoneaza_foi.curata_legenda(
        wb, total_fluxuri=len(O.ORDINE), total_corelatii=12 + len(CORELATII))

    istoric.construieste(wb, mutate=mutate, orfane=orfane,
                         reformulari=dreform.INLOCUIRI,
                         resturi=resturi, absorbite=absorbite, statusuri=statusuri)

    schimbate = renumeroteaza.in_workbook(wb, O.HARTA, exclude=("Istoric",))
    ramase = renumeroteaza.ramase(wb, set(O.HARTA), exclude=("Istoric",))
    if ramase:
        raise SystemExit(f"ID-uri vechi rămase după renumerotare: "
                         f"{ {k: v[:3] for k, v in ramase.items()} }")

    # Ancorele se pun DUPĂ renumerotare, ca să poarte ID-urile finale.
    seed = openpyxl.load_workbook(SEED)
    harta_fm = ancore.harta_flux_modul(dmodule.MODULE, seed["Index module"], O.HARTA)
    anc = ancore.adauga(wb, harta_fm)

    # foaia de întrebări + marcajele ❓ pe fluxurile și corelațiile atinse
    marcaje = foaie_intrebari.construieste(wb)
    anc["intrebari"] = foaie_intrebari.marcheaza(wb, marcaje)

    return orfane, dupl, schimbate, anc


def main():
    if not os.path.exists(SEED):
        raise SystemExit(f"Lipsește fișierul-sămânță: {SEED}")
    wb = openpyxl.load_workbook(SEED)
    extinde_plan_conturi(wb)
    extinde_analitice(wb)
    extinde_fluxuri(wb)
    extinde_corelatii(wb)
    extinde_matrice(wb)
    actualizeaza_legenda(wb)
    actualizeaza_index(wb)
    orfane, dupl, schimbate, anc = ordoneaza_canonic(wb)

    os.makedirs(os.path.dirname(IESIRE), exist_ok=True)
    wb.save(IESIRE)
    recalc(IESIRE)
    print(f"scris: {os.path.relpath(IESIRE, RADACINA)}\n"
          f"   {len(O.ORDINE)} fluxuri ordonate pe clase, {schimbate} celule renumerotate\n"
          f"   {dupl} rânduri duplicate contopite, {len(orfane)} rânduri orfane "
          f"mutate în Istoric\n"
          f"   ancore de modul: {anc['fluxuri']} fluxuri, {anc['corelatii']} corelații, "
          f"{anc['matrice']} rânduri de matrice\n"
          f"   foaia Întrebări deschise + {anc['intrebari']} marcaje ❓")


if __name__ == "__main__":
    main()

"""Generează dist/Module_Declarative_Fluxuri.xlsx.

Extindere pe bază de seed: se încarcă workbook-ul original din
`surse/training-4-2026-08-14/` și se adaugă modulele noi din `date/module/`,
fiecare cu setul complet de patru foi, pe tiparul lui MOD_INCHIDERE_TVA:

    Declarații_<COD>   input galben (font albastru) + calcule auto
    Reguli_<COD>       tabele fixe — se editează doar când se schimbă legea
    Jurnale_<COD>      blocuri de înregistrări, prin formulă, cu celule Check
    NotaExport_<COD>   1 rând = 1 înregistrare, cu coloana Include

`Parametri` primește pragurile noi ÎNAINTE de construirea modulelor, ca acestea
să poată referi celulele reale (nu coordonate ghicite).

Rulare:  python build/build_module.py
"""
import os
import sys

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from build import stil  # noqa: E402
from build.foaie import Foaie  # noqa: E402
from build.recalc import recalc  # noqa: E402
from date.module import MODULE, PARAMETRI_NOI  # noqa: E402
from date.module.comun import sufix as sufix_modul  # noqa: E402
from date import ordine as O  # noqa: E402
from build import renumeroteaza  # noqa: E402

RADACINA = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SEED = os.path.join(RADACINA, "surse", "training-4-2026-08-14", "Module_Declarative_Fluxuri.xlsx")
IESIRE = os.path.join(RADACINA, "dist", "Module_Declarative_Fluxuri.xlsx")


def extinde_parametri(wb):
    """Scrie parametrii globali și le dă NUME definite. Întoarce {cheie: nume}.

    Modulele referă `P["prag_mf"]`, deci formula scrisă în fișier devine
    `=param_prag_mf` în loc de `=Parametri!B21`. Două câștiguri:

    - se citește în Excel fără să sari la altă foaie ca să afli ce e B21;
    - dacă cineva inserează un rând în `Parametri`, numele urmează celula, iar toate
      cele 13 module rămân corecte.

    E convenția din templateul extern de deconturi (`decl_cota_standard`), adusă aici.
    """
    ws = wb["Parametri"]
    rand = ws.max_row + 2
    stil.scrie(ws, rand, 1, "Parametri globali — referiți prin nume definit (param_*)",
               font=stil.F_TITLU_BLOC)
    rand += 1
    stil.scrie(ws, rand, 1, "Numele din coloana C se folosește direct în formule. "
                            "Schimbă valoarea aici o singură dată; toate modulele o preiau.",
               font=stil.F_NOTA, align=stil.A_WRAP)
    rand += 1
    if ws.column_dimensions["C"].width is None or ws.column_dimensions["C"].width < 50:
        ws.column_dimensions["C"].width = 56

    stil.scrie(ws, rand, 1, "Parametru", font=stil.F_CAP_TABEL)
    stil.scrie(ws, rand, 2, "Valoare", font=stil.F_CAP_TABEL)
    stil.scrie(ws, rand, 3, "Nume definit · temei", font=stil.F_CAP_TABEL)
    rand += 1

    refs = {}
    for cheie, eticheta, valoare, nota in PARAMETRI_NOI:
        nume = f"param_{cheie}"
        stil.scrie(ws, rand, 1, eticheta, font=stil.F_NORMAL, align=stil.A_WRAP)
        stil.scrie(ws, rand, 2, valoare, font=stil.F_INPUT, fill=stil.FILL_INPUT)
        stil.scrie(ws, rand, 3, f"{nume} — {nota}", font=stil.F_NOTA, align=stil.A_WRAP)
        wb.defined_names.add(DefinedName(nume, attr_text=f"Parametri!$B${rand}"))
        refs[cheie] = nume
        rand += 1
    return refs


def insereaza_randuri(ws, rand, n):
    """`insert_rows` care deplasează ȘI celulele îmbinate.

    openpyxl mută valorile la inserare, dar lasă `merged_cells` pe loc. Pe o foaie
    cu îmbinări (CatalogModule are A1:H1, A17:H17 …) rândurile noi cad peste un
    merge rămas nedeplasat, iar tot ce se scrie în afara colțului stânga-sus se
    pierde tăcut la salvare. De aceea îmbinările se desfac, se inserează, apoi se
    refac deplasate.
    """
    afectate = [m for m in list(ws.merged_cells.ranges) if m.min_row >= rand]
    for m in afectate:
        ws.unmerge_cells(str(m))
    ws.insert_rows(rand, amount=n)
    for m in afectate:
        ws.merge_cells(start_row=m.min_row + n, start_column=m.min_col,
                       end_row=m.max_row + n, end_column=m.max_col)


def extinde_catalog(wb, module):
    ws = wb["CatalogModule"]
    rand = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().startswith("REGULĂ DE SUMONARE"):
            rand = r
            break
    if rand is None:
        raise SystemExit("Nu găsesc blocul „REGULĂ DE SUMONARE” în CatalogModule")

    # Un modul care exista deja în catalog (MOD_LEASING_FIN era „EXEMPLU EXTERN”) se
    # ACTUALIZEAZĂ în loc. Altfel catalogul ajunge să aibă două rânduri pentru același
    # modul, care se contrazic — exact defectul semnalat la citirea integrală.
    existente = {}
    for r in range(1, ws.max_row + 1):
        cod = str(ws.cell(row=r, column=2).value or "").strip()
        if cod.startswith("MOD_"):
            existente[cod] = r

    noi = [m for m in module if m.COD not in existente]
    if noi:
        insereaza_randuri(ws, rand, len(noi) + 1)
        # rândurile de mai jos s-au deplasat
        existente = {c: (r + len(noi) + 1 if r >= rand else r) for c, r in existente.items()}

    # Coloana `Activ` se scrie din `CATALOG['activ']`, pentru TOATE modulele — și cele
    # care aveau deja rând în sămânță. Până acum steagul rămânea cum îl lăsase sămânța,
    # iar împărțirea DA/NU ajunsese să însemne „livrat înainte / după etapa 4”: DA pe
    # cele șapte moștenite, NU pe tot ce fusese portat, inclusiv pe MOD_INCHIDERE_LUNARA,
    # singurul care chiar se rulează în fiecare lună. Acum DA înseamnă cadență lunară
    # NECONDIȚIONATĂ — fără contract, activ sau eveniment declanșator.
    i = 0
    for m in module:
        cat = m.CATALOG
        if m.COD in existente:
            r = existente[m.COD]
        else:
            r = rand + i
            i += 1
        activ = cat.get("activ", "NU")
        stil.scrie(ws, r, 1, activ, font=stil.F_INPUT,
                   fill=stil.FILL_DA if activ == "DA" else stil.FILL_NU,
                   align=stil.A_CENTER)
        for col, val in enumerate([m.COD, cat["fluxuri"], cat["tip"], cat["variabile"],
                                   cat["porti"], cat["blocuri"], "IMPLEMENTAT"], start=2):
            stil.scrie(ws, r, col, val, font=stil.F_NORMAL, fill=stil.FILL_DA,
                       align=stil.A_WRAP)


#: Prefixele celor patru foi ale unui modul. `Verificări_`/`Abateri_` sunt varianta
#: modulelor care dau un verdict, nu o înregistrare (MOD_INCHIDERE_LUNARA).
PREFIXE = ("Declarații_", "Reguli_", "Jurnale_", "NotaExport_", "Verificări_",
           "Abateri_", "Registru_")


def scoate_foile_de_samanta(wb, module):
    """Șterge foile pe care un modul portat urmează să le regenereze.

    Un modul care exista DOAR ca foi de sămânță (MOD_INTERMEDIAR și celelalte șase) își
    primește acum foile din cod. Fără ștergerea asta, `Foaie.__init__` ridică „foaia
    există deja”.

    Ce apără pasul ăsta: poarta de conservare cere ca fiecare linie din sămânță să se
    regăsească în fișierul generat. Deci dacă portarea uită o propoziție din foaia
    veche, poarta 9 o numește. Portarea se verifică singură — nu pe încrederea că am
    transcris tot, ci pe mulțimea de text.
    """
    coduri = {sufix_modul(m) for m in module}
    scoase = [s for s in list(wb.sheetnames)
              if any(s.startswith(p) and s.removeprefix(p) in coduri for p in PREFIXE)]
    for s in scoase:
        del wb[s]
    return scoase


def rescrie_instructiuni(wb, module):
    """Refac foaia `Instructiuni` din catalog. Întoarce rândurile vechi, pentru declarare.

    Foaia era încă scrisă pentru un fișier cu UN modul: „Cum se folosește
    MOD_INCHIDERE_TVA (primul modul implementat)”, cu SALARII / DECONT / LEASING_FIN
    date drept „EXEMPLU EXTERN” când sunt interne de mult. Descria tiparul ca
    `Declarații → Jurnale → NotaExport`, fără `Reguli`, și nu pomenea deloc varianta
    `Verificări_` / `Abateri_`.

    Avea și un defect viu: lista „Fluxuri corectate” ajunsese „F-311,06,07,14,15,…” —
    renumerotarea potrivește token-uri întregi, deci a prefăcut doar primul element al
    unei enumerări scrise prescurtat, lăsând restul în numerotarea veche. O listă
    jumătate-tradusă e mai rea decât una netradusă: arată actualizată.
    """
    ws = wb["Instructiuni"]
    vechi = [str(c.value).strip() for r in ws.iter_rows() for c in r
             if isinstance(c.value, str) and c.value.strip()]

    lunare = [m for m in module if m.CATALOG.get("activ") == "DA"]
    linii = [
        ("MODULE DECLARATIVE — instrucțiuni de utilizare", stil.F_TITLU),
        ("", None),
        ("Ce este acest fișier", stil.F_TITLU_BLOC),
        (f"Cele {len(module)} module declarative care execută fluxurile din "
         "„Plan de conturi”. Fluxul explică de ce; modulul face înregistrarea. "
         "Completezi câteva variabile într-o foaie de input și ies jurnalele și nota "
         "de export.", stil.F_NORMAL),
        ("Toate foile se generează din depozit, la fiecare build. Nu se editează "
         "structura aici: o modificare făcută în fișier se pierde la următorul `make`.",
         stil.F_NOTA),
        ("", None),
        ("Cum se folosește ORICE modul", stil.F_TITLU_BLOC),
        ("1. CatalogModule — pune Activ=DA pe modulele de care ai nevoie luna asta.",
         stil.F_NORMAL),
        ("2. Declarații_<MODUL> — completează doar celulele galbene cu scris albastru. "
         "Restul sunt formule.", stil.F_NORMAL),
        ("3. Reguli_<MODUL> — nu se completează. E regula contabilă sau fiscală, cu "
         "temeiul ei, plus ce anume rupe corelațiile. Se citește când rezultatul te "
         "surprinde.", stil.F_NORMAL),
        ("4. Jurnale_<MODUL> — verifică fiecare celulă Check. Trebuie OK înainte de a "
         "lua rezultatul de bun.", stil.F_NORMAL),
        ("5. NotaExport_<MODUL> — filtrează Include=DA și înregistrează în program.",
         stil.F_NORMAL),
        ("", None),
        ("Două tipare de foi", stil.F_TITLU_BLOC),
        ("Declarații → Reguli → Jurnale → NotaExport — tiparul obișnuit, pentru modulele "
         "care produc înregistrări.", stil.F_NORMAL),
        ("Declarații → Reguli → Verificări → Abateri — pentru modulul care produce un "
         "VERDICT, nu o înregistrare (MOD_INCHIDERE_LUNARA): confruntă balanța cu "
         "corelațiile lunii și listează doar ce nu se potrivește.", stil.F_NORMAL),
        ("MOD_DECONT are în plus un Registru_, pentru liniile decontului.", stil.F_NOTA),
        ("", None),
        ("Ce înseamnă Activ=DA", stil.F_TITLU_BLOC),
        ("Cadență lunară NECONDIȚIONATĂ: modulul se rulează în fiecare lună la o firmă "
         "obișnuită, fără să fie nevoie de un contract, un activ sau un eveniment "
         "declanșator. Restul se activează când ai cazul pe masă.", stil.F_NORMAL),
        (f"Implicit DA: {', '.join(m.COD for m in lunare)}.", stil.F_NORMAL),
        ("", None),
        ("Principiu de design", stil.F_TITLU_BLOC),
        ("Modulul nu inventează conturi. Folosește doar conturile din politica declarată "
         "+ rolurile din Plan de conturi. Porțile de calitate (ΣDr=ΣCr, stare terminală, "
         "pas revelator) rămân aceleași ca în foaia Fluxuri.", stil.F_NORMAL),
    ]

    for r in range(ws.max_row, 0, -1):
        ws.delete_rows(r)
    ws.column_dimensions["A"].width = 110
    for i, (text, font) in enumerate(linii, start=1):
        if text:
            stil.scrie(ws, i, 1, text, font=font, align=stil.A_WRAP_TOP)
    return vechi


def main():
    if not os.path.exists(SEED):
        raise SystemExit(f"Lipsește fișierul-sămânță: {SEED}")
    wb = openpyxl.load_workbook(SEED)

    scoase = scoate_foile_de_samanta(wb, MODULE)
    P = extinde_parametri(wb)

    def fabrica(nume, latimi=None):
        return Foaie(wb, nume, latimi)

    for m in MODULE:
        m.construieste(fabrica, P)

    extinde_catalog(wb, MODULE)
    rescrie_instructiuni(wb, MODULE)

    # `date/` foloseste numerotarea VECHE a fluxurilor; renumerotarea e o singura
    # trecere finala, ca in workbook-ul de plan. Fara ea, catalogul de module ar arata
    # F-46 langa F-107 -- exact genul de drift pe care sistemul trebuie sa-l excluda.
    schimbate = renumeroteaza.in_workbook(wb, O.HARTA)
    ramase = renumeroteaza.ramase(wb, set(O.HARTA))
    if ramase:
        raise SystemExit(f"ID-uri vechi ramase in modulele declarative: "
                         f"{ {k: v[:3] for k, v in ramase.items()} }")

    os.makedirs(os.path.dirname(IESIRE), exist_ok=True)
    wb.save(IESIRE)
    recalc(IESIRE)
    print(f"scris: {os.path.relpath(IESIRE, RADACINA)}  "
          f"({len(MODULE)} module, {schimbate} celule renumerotate"
          + (f", {len(scoase)} foi de sămânță regenerate din cod" if scoase else "")
          + ")")


if __name__ == "__main__":
    main()

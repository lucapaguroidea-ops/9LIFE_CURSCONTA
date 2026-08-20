"""Generează dist/parcurs-training-nou.md.

Documentul are două straturi. Cel scris de mână stă în `date/parcurs.py` — întrebările
de judecată, punctele de convergență, traseul real. Cel generat e aici: tot ce enumeră
starea sistemului.

Distincția e regula centrală a documentului. O listă de porți sau de module scrisă de
mână începe să diveargă din prima zi, iar un document de îndrumare care minte despre
starea sistemului e mai rău decât niciunul.

De remarcat: „harta documentelor de referință” nu e scrisă nicăieri — se CITEȘTE din
tabelul de structură al foii `Legendă`, pe care poarta 14 îl ține complet. Deci nici
măcar harta nu e un al doilea adevăr.

Rulare:  python build/parcurs.py   (după `make build`)
"""
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from date import documente as ddoc  # noqa: E402
from date import ordine as O  # noqa: E402
from date import parcurs as P  # noqa: E402

RADACINA = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PLAN = os.path.join(RADACINA, "dist", "Plan_de_conturi_ROL_Analitice_Fluxuri_SAGA.xlsx")
MODULE = os.path.join(RADACINA, "dist", "Module_Declarative_Fluxuri.xlsx")
VERIFICA = os.path.join(RADACINA, "build", "verifica.py")
IESIRE = os.path.join(RADACINA, "dist", "parcurs-training-nou.md")

RE_POARTA = re.compile(r"^\s*(\d+)\.\s+(.+)$")


def porti():
    """[(număr, descriere)] — din lista canonică din docstring-ul lui verifica.py.

    Descrierile NU se extrag din mesajele `ok(...)`: acelea sunt f-string-uri, iar
    tăierea la prima interpolare producea fraze trunchiate („toate cele”). Lista din
    docstring e scrisă pentru citit, iar poarta 15 o ține sincronizată cu porțile
    chemate efectiv.
    """
    with open(VERIFICA, encoding="utf-8") as f:
        doc = f.read().split('"""')[1]
    out, curent = [], None
    for linie in doc.split("\n"):
        m = RE_POARTA.match(linie)
        if m:
            curent = [int(m.group(1)), m.group(2).strip()]
            out.append(curent)
        elif curent and linie.startswith("     ") and linie.strip():
            curent[1] += " " + linie.strip()
        elif not linie.strip():
            curent = None
    return [(n, [d]) for n, d in out]


def harta_referinta():
    """Tabelul de structură din foaia Legendă — citit, nu rescris."""
    ws = openpyxl.load_workbook(PLAN)["Legendă"]
    randuri, inauntru = [], False
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(row=r, column=1).value or "").strip()
        b = str(ws.cell(row=r, column=2).value or "").strip()
        if a.startswith("5. STRUCTURA"):
            inauntru = True
            continue
        if inauntru and a.startswith("6. "):
            break
        if inauntru and a and a != "Foaie":
            randuri.append((a, b))
    return randuri


def module():
    ws = openpyxl.load_workbook(MODULE)["CatalogModule"]
    out = []
    for r in range(1, ws.max_row + 1):
        cod = str(ws.cell(row=r, column=2).value or "").strip()
        if cod.startswith("MOD_"):
            out.append((cod, str(ws.cell(row=r, column=3).value or "").strip()))
    return out


def fisiere_continut():
    """Fișierele din `date/` care țin conținut, cu prima linie din docstring."""
    out = []
    baza = os.path.join(RADACINA, "date")
    for nume in sorted(os.listdir(baza)):
        if not nume.endswith(".py") or nume in ("__init__.py", "comun.py"):
            continue
        with open(os.path.join(baza, nume), encoding="utf-8") as f:
            prima = ""
            for linie in f:
                if linie.strip().startswith('"""'):
                    prima = linie.strip().strip('"').strip()
                    break
        out.append((f"date/{nume}", prima))
    return out


def _tabel(capete, randuri):
    out = ["| " + " | ".join(capete) + " |",
           "|" + "|".join(["---"] * len(capete)) + "|"]
    out += ["| " + " | ".join(str(c) for c in r) + " |" for r in randuri]
    return out


def construieste():
    L = ["# Parcursul unui set nou de notițe", "",
         "*Document de îndrumare. Nu explică sistemul — îl indică.*", "", "---", "",
         "## 0. Cum se folosește", "", P.INTRO, "",
         "> **Testul aplicat fiecărui paragraf:** dacă documentul de referință s-ar "
         "schimba mâine, paragraful ăsta ar deveni fals? Dacă da, e parafrază, nu "
         "îndrumare — și nu are ce căuta aici.", "",
         "Secțiunile marcate *[generat]* se recitesc din cod și din workbook-uri la "
         "fiecare `make tot`. Nu le edita: se suprascriu.", "", "---", ""]

    # 1 ------------------------------------------------------------------
    L += ["## 1. Unde se uită *[generat]*", "",
          "Tabelul de mai jos nu e scris aici: e citit din foaia `Legendă`, secțiunea "
          "„STRUCTURA FINALĂ A WORKBOOK-ULUI”, pe care poarta 14 o ține completă. Dacă "
          "apare o foaie nouă în workbook, apare și aici.", ""]
    L += _tabel(["Foaie", "Ce ține"], harta_referinta())
    L += ["", "**Documente conexe:** cele trei `.md`/`.docx` revizuite (structura lor "
          "canonică e în `date/documente.py`) și `intrebari-formator.md` — lista "
          "întrebărilor deschise, aceeași sursă cu foaia `Întrebări deschise`.", "",
          "---", ""]

    # 2 ------------------------------------------------------------------
    L += ["## 2. Faza A — notițe brute → document revizuit", "", P.FAZA_A_INTRO, ""]
    for i, (intrebare, dece, unde) in enumerate(P.FAZA_A, start=1):
        L += [f"**{i}. {intrebare}**", "", dece, "", f"↳ {unde}", ""]
    L += ["### Forma rezultatului *[generat]*", "",
          "Documentul revizuit are aceeași legendă de marcaje ca celelalte trei și "
          "anexele denumite canonic. Poarta 13 verifică asta.", ""]
    L += _tabel(["Anexa", "Conținut"],
                [(f"**{k}**", v) for k, v in ddoc.ANEXE.items()])
    L += ["", "Nu orice document are toate anexele — le are pe cele pentru care există "
          "conținut real. Vezi `date/documente.py` pentru ce anexă e servită de ce "
          "secțiune în fiecare document.", "", "---", ""]

    # 3 ------------------------------------------------------------------
    L += ["## 3. Faza B — document revizuit → Excel", "", P.FAZA_B_INTRO, ""]
    for titlu, corp in P.RUNBOOK:
        L += [f"**{titlu}**", "", corp, ""]

    L += ["### Unde se scrie fiecare lucru *[generat]*", ""]
    L += _tabel(["Fișier", "Ce ține"],
                [(f"`{f}`", d) for f, d in fisiere_continut()])
    L += ["", "### Clasele de fluxuri *[generat]*", "",
          "ID-ul codifică clasa contului principal. Un flux nou primește următorul "
          "număr liber din clasa lui — `ordine.urmatorul_liber(clasa)` — și stă fizic "
          "la locul lui.", ""]
    L += _tabel(["Bloc", "Clasa", "Fluxuri acum", "Următorul liber"],
                [(f"`F-{cls}xx`", titlu.split("—")[1].strip() if "—" in titlu else titlu,
                  len(bloc), f"`{O.urmatorul_liber(cls)}`")
                 for cls, titlu, bloc in O.BLOCURI])

    L += ["", "### Modulele existente *[generat]*", "",
          "Un flux e util și fără modul: fluxul explică, modulul execută. Dacă adaugi "
          "un modul, declară fluxurile în `CATALOG['fluxuri']` — de acolo se derivă "
          "toate ancorele.", ""]
    L += _tabel(["Modul", "Fluxuri acoperite"],
                [(f"`{c}`", f) for c, f in module()])

    L += ["", "### Ce verifică fiecare poartă *[generat]*", ""]
    L += _tabel(["Poarta", "Verifică"],
                [(f"**{n}**", " · ".join(m)) for n, m in porti()])
    L += ["", "Motivul fiecărei porți e scris în `build/verifica.py`, lângă ea. Când o "
          "poartă pică, acolo scrie de ce există.", "", "---", ""]

    # 4 ------------------------------------------------------------------
    L += ["## 4. Puncte de convergență", "", P.CONVERGENTA_INTRO, ""]
    L += _tabel(["Ce se poate rupe", "Ce se pierde", "Prins de"],
                [(p, ce, f"poarta {g}" if g else "**nimic — verifică tu**")
                 for p, ce, g in P.CONVERGENTA])
    L += ["", "---", ""]

    # 5 ------------------------------------------------------------------
    L += ["## 5. Traseul trainingului 3, pas cu pas", "", P.TRASEU_INTRO, ""]
    for i, (pas, nota) in enumerate(P.TRASEU, start=1):
        L += [f"**{i}.** {pas}", "", f"   {nota}", ""]
    L += ["---", ""]

    # 6 ------------------------------------------------------------------
    L += ["## 6. Goluri cunoscute", "",
          "Lucruri care ar merita mecanizate și care azi cad în sarcina cititorului. "
          "Enumerate ca să nu fie confundate cu ceva acoperit.", ""]
    for titlu, corp in P.GOLURI:
        L += [f"**{titlu}**", "", corp, ""]

    L += ["---", "",
          f"*Secțiunile [generat] provin din `build/verifica.py`, `date/ordine.py`, "
          f"`date/documente.py` și din workbook-urile construite. "
          f"{len(porti())} porți, {len(module())} module, "
          f"{len(O.ORDINE)} fluxuri la data generării.*", ""]
    return "\n".join(L)


def main():
    for cale in (PLAN, MODULE):
        if not os.path.exists(cale):
            raise SystemExit(f"Rulează întâi `make build` — lipsește {os.path.basename(cale)}")
    text = construieste()
    os.makedirs(os.path.dirname(IESIRE), exist_ok=True)
    with open(IESIRE, "w", encoding="utf-8") as f:
        f.write(text)
    print(f"scris: {os.path.relpath(IESIRE, RADACINA)}  "
          f"({len(porti())} porți, {len(module())} module, "
          f"{len(P.CONVERGENTA)} puncte de convergență)")


if __name__ == "__main__":
    main()

"""Paleta de stiluri a sistemului — extrasă din workbook-urile originale (training 4).

Scop: conținutul nou adăugat pe foile existente trebuie să fie indistinctibil de
original. Culorile de mai jos sunt exact cele folosite în
`Plan_de_conturi_ROL_Analitice_Fluxuri_SAGA.xlsx` și `Module_Declarative_Fluxuri.xlsx`.

Convenție openpyxl: culorile se dau ca RRGGBB (fără alfa). La citire openpyxl le
raportează prefixate cu 00/FF — e același lucru.
"""
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ---------------------------------------------------------------- culori (RRGGBB)
BLEUMARIN = "1F4E79"   # antet clasă / titluri
ALB = "FFFFFF"
GRUP = "D6DCE4"        # rânduri de grupă (cont din 2 cifre) în Plan de conturi
PORTOCALIU = "FCE4D6"  # rând de pas normal în Fluxuri
VERDE = "E2EFDA"       # rând de pas revelator în Fluxuri
VERDE_TARE = "C6EFCE"  # catalog: flux didactic / catalog module: Activ=DA
BLEU = "DDEBF7"        # catalog: flux ne-didactic
GALBEN = "FFF2CC"      # celulă de input (module declarative)
ALBASTRU = "0000FF"    # font pe celulă de input
GRI = "666666"         # text de notă / „Principiul:"
ROSU = "C00000"        # avertismente (★, storno de roșu)

# ------------------------------------------------------------------------- fonturi
def font(bold=False, color=None, size=9, italic=False, name="Arial"):
    return Font(name=name, bold=bold, italic=italic, size=size, color=color)

F_ANTET = font(bold=True, color=ALB, size=10)
F_TITLU = font(bold=True, color=BLEUMARIN, size=14)
F_TITLU_BLOC = font(bold=True, color=BLEUMARIN, size=11)
F_SUBTITLU = font(bold=True, color=BLEUMARIN, size=13)
F_GRUP = font(bold=True, size=9)
F_CAP_TABEL = font(bold=True, size=10)
F_CAP_TABEL_ALB = font(bold=True, color=ALB, size=11)
F_NORMAL = font(size=9)
F_NOTA = font(italic=True, color=GRI, size=9)
F_INPUT = font(color=ALBASTRU, size=10)
F_AVERTISMENT = font(color=ROSU, size=9)

# --------------------------------------------------------------------------- fill
def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

FILL_ANTET = fill(BLEUMARIN)
FILL_GRUP = fill(GRUP)
FILL_PAS = fill(PORTOCALIU)
FILL_PAS_REVELATOR = fill(VERDE)
FILL_DIDACTIC = fill(VERDE_TARE)
FILL_NEDIDACTIC = fill(BLEU)
FILL_INPUT = fill(GALBEN)
FILL_DA = fill(VERDE_TARE)
FILL_NU = fill(GALBEN)

# --------------------------------------------------------------------- aliniere
A_WRAP = Alignment(wrap_text=True, vertical="center")
A_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
A_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

_thin = Side(style="thin", color="BFBFBF")
BORDER_SUBTIRE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def scrie(ws, row, col, value, *, font=None, fill=None, align=None, border=None,
          text=False):
    """Scrie o celulă cu stil, întorcând obiectul celulă.

    `text=True` forțează celula să fie TEXT chiar dacă valoarea începe cu „=”.

    E nevoie de asta oriunde se documentează o formulă. Foaia `Istoric` listează
    înlocuirile declarate din `date/reformulari.py`, iar de când portarea modulelor a
    adus acolo formule („=IF(CatalogModule!A13=…)”), openpyxl le scria drept FORMULE
    live — care apoi refereau o foaie inexistentă în workbook-ul de plan. Motorul de
    recalcul se plângea, dar build-ul mergea mai departe, iar celula rămânea stricată.
    """
    c = ws.cell(row=row, column=col, value=value)
    if text and isinstance(value, str) and value.startswith("="):
        c.data_type = "s"
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if align is not None:
        c.alignment = align
    if border is not None:
        c.border = border
    return c
